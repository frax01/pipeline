import requests
import time
from openpyxl import load_workbook
from tqdm import tqdm

MCP_SDKS = {
    "@modelcontextprotocol/sdk",
    "fastmcp",
    "modelcontextprotocol",
    "mcp"
}

def check_mcp_server(pkg: str):
    """
    Ritorna:
    (status, version, maintainers, repository)
    """
    url = f"https://registry.npmjs.org/{pkg}"
    r = requests.get(url, timeout=10)
    if r.status_code != 200:
        return ("ERROR", None, None, None)

    data = r.json()
    latest = data["dist-tags"]["latest"]
    version_data = data["versions"][latest]

    deps = version_data.get("dependencies", {})
    dev_deps = version_data.get("devDependencies", {})
    all_deps = set(deps) | set(dev_deps)

    is_mcp = any(dep in MCP_SDKS for dep in all_deps)

    maintainers = data.get("maintainers", [])
    maintainers_str = ", ".join(
        m.get("name", "") for m in maintainers if isinstance(m, dict)
    )

    repo = data.get("repository", {})
    if isinstance(repo, dict):
        repo_url = repo.get("url", "")
    else:
        repo_url = str(repo)
    
    # normalizzazione URL repository
    if repo_url:
        if repo_url.startswith("git+"):
            repo_url = repo_url[4:]
        if repo_url.endswith(".git"):
            repo_url = repo_url[:-4]


    return (
        "MCP" if is_mcp else "NO",
        latest,
        maintainers_str,
        repo_url
    )


# ======================
# 1. Carica Excel
# ======================
wb = load_workbook("mcp_packages.xlsx")
ws_all = wb["mcp_packages"]

if "real_mcp_servers" not in wb.sheetnames:
    ws_real = wb.create_sheet("real_mcp_servers")
    ws_real.append([
        "package_name",
        "status",
        "version",
        "maintainers",
        "repository"
    ])
else:
    ws_real = wb["real_mcp_servers"]

# ======================
# 2. Stato già processato
# ======================
processed = {
    ws_real[f"A{r}"].value
    for r in range(2, ws_real.max_row + 1)
    if ws_real[f"A{r}"].value
}

packages = [
    ws_all[f"A{r}"].value
    for r in range(2, ws_all.max_row + 1)
    if ws_all[f"A{r}"].value and ws_all[f"A{r}"].value not in processed
]

print(f"Pacchetti totali       : {ws_all.max_row - 1}")
print(f"Già processati         : {len(processed)}")
print(f"Rimanenti da analizzare: {len(packages)}")

# ======================
# 3. Analisi stop & resume
# ======================
try:
    for pkg in tqdm(packages, desc="Verifica MCP server", unit="pkg"):
        try:
            status, version, maintainers, repo = check_mcp_server(pkg)
        except Exception:
            status, version, maintainers, repo = ("ERROR", None, None, None)

        ws_real.append([
            pkg,
            status,
            version,
            maintainers,
            repo
        ])

        # salvataggio incrementale (fondamentale)
        wb.save("mcp_packages.xlsx")

        time.sleep(0.05)

except KeyboardInterrupt:
    print("\n⏸ Interrotto dall'utente. Stato salvato.")

# ======================
# 4. Salvataggio finale
# ======================
wb.save("mcp_packages.xlsx")
print("✅ Operazione completata / salvata correttamente")