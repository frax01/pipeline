"""
STEP 3: Unione dei risultati dalle 9 VM e rilevamento duplicati.

Questo script:
  1. Legge tutti i file vm_X_results.xlsx dalla cartella vm_results/
  2. Unisce tutti gli hash-URL in un'unica mappa
  3. Rileva i duplicati (stesso hash = stesso contenuto)
  4. Salva:
     - duplicates.json: tutti i duplicati (original_url + duplicate_url + hash)
     - servers_no_duplicates.xlsx: Excel pulito senza URL duplicati
     - hash_url_mapping.xlsx: Excel con hash + URL per analisi incrementali future

Da lanciare DOPO aver copiato tutti i vm_X_results.xlsx nella cartella vm_results/.
"""

import os
import sys
import json
import openpyxl
import openpyxl.styles

# === CONFIGURAZIONE ===
SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
INPUT_FILE = os.path.join(SCRIPT_DIR, "0.0. All servers without duplicates, .git and _ (66333).xlsx")
RESULTS_DIR = os.path.join(SCRIPT_DIR, "vm_results")
DUPLICATES_FILE = os.path.join(SCRIPT_DIR, "duplicates.json")
OUTPUT_CLEAN_EXCEL = os.path.join(SCRIPT_DIR, "servers_no_duplicates.xlsx")
OUTPUT_HASH_EXCEL = os.path.join(SCRIPT_DIR, "hash_url_mapping.xlsx")


def read_original_urls(filepath):
    """Legge gli URL originali per mantenere l'ordine."""
    wb = openpyxl.load_workbook(filepath, read_only=True, data_only=True)
    ws = wb.active
    urls = []
    for row in ws.iter_rows(min_row=2, min_col=1, max_col=1, values_only=True):
        val = row[0]
        if val is not None:
            url = str(val).strip()
            if url:
                urls.append(url)
    wb.close()
    return urls


def read_vm_results(filepath):
    """
    Legge un file risultati VM.
    Colonna A = URL, Colonna B = hash (o ERROR:...).
    Ritorna lista di (url, hash).
    """
    wb = openpyxl.load_workbook(filepath, read_only=True, data_only=True)
    ws = wb.active
    results = []
    for row in ws.iter_rows(min_row=2, min_col=1, max_col=2, values_only=True):
        url = row[0]
        hash_val = row[1] if len(row) > 1 else None
        if url is not None and hash_val is not None:
            results.append((str(url).strip(), str(hash_val).strip()))
    wb.close()
    return results


def save_clean_excel(urls, filepath):
    """Salva Excel con gli URL senza duplicati."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Servers (no duplicates)"
    header_font = openpyxl.styles.Font(bold=True, size=12)

    ws["A1"] = "GitHub URL"
    ws["A1"].font = header_font
    ws.column_dimensions["A"].width = 80

    for i, url in enumerate(urls, start=2):
        ws.cell(row=i, column=1, value=url)

    wb.save(filepath)


def save_hash_mapping_excel(hash_url_pairs, filepath):
    """Salva Excel con hash (colonna A) e URL (colonna B)."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Hash-URL Mapping"
    header_font = openpyxl.styles.Font(bold=True, size=12)

    ws["A1"] = "SHA-256 Hash"
    ws["A1"].font = header_font
    ws["B1"] = "GitHub URL"
    ws["B1"].font = header_font
    ws.column_dimensions["A"].width = 70
    ws.column_dimensions["B"].width = 80

    for i, (hash_val, url) in enumerate(hash_url_pairs, start=2):
        ws.cell(row=i, column=1, value=hash_val)
        ws.cell(row=i, column=2, value=url)

    wb.save(filepath)


def main():
    print("=" * 70)
    print("  MERGE VM RESULTS - Unione risultati e rilevamento duplicati")
    print("=" * 70)
    print()

    # Verifica cartella risultati
    if not os.path.exists(RESULTS_DIR):
        print(f"ERRORE: Cartella risultati non trovata: {RESULTS_DIR}")
        print(f"  Crea la cartella e copia i file vm_X_results.xlsx al suo interno.")
        sys.exit(1)

    # Trova tutti i file risultati
    result_files = sorted([
        f for f in os.listdir(RESULTS_DIR)
        if f.endswith("_results.xlsx") and f.startswith("vm_")
    ])

    if not result_files:
        print(f"ERRORE: Nessun file vm_X_results.xlsx trovato in {RESULTS_DIR}")
        print(f"  Assicurati di aver copiato i risultati dalle VM.")
        sys.exit(1)

    print(f"[1/5] File risultati trovati: {len(result_files)}")
    for f in result_files:
        print(f"  - {f}")
    print()

    # Leggi ordine originale URL
    print("[2/5] Lettura ordine originale degli URL...")
    original_urls = read_original_urls(INPUT_FILE)
    print(f"  URL originali: {len(original_urls)}")
    print()

    # Leggi tutti i risultati VM
    print("[3/5] Lettura risultati da tutte le VM...")
    all_results = {}  # url -> hash
    errors = {}  # url -> error_message
    total_read = 0

    for filename in result_files:
        filepath = os.path.join(RESULTS_DIR, filename)
        results = read_vm_results(filepath)
        total_read += len(results)

        success_count = 0
        error_count = 0
        for url, hash_val in results:
            if hash_val.startswith("ERROR:"):
                errors[url] = hash_val
                error_count += 1
            else:
                all_results[url] = hash_val
                success_count += 1

        print(f"  {filename}: {len(results)} righe ({success_count} ok, {error_count} errori)")

    print(f"\n  Totale letti: {total_read}")
    print(f"  Con hash valido: {len(all_results)}")
    print(f"  Con errore: {len(errors)}")

    # Verifica copertura
    original_set = set(original_urls)
    analyzed_set = set(all_results.keys()) | set(errors.keys())
    missing = original_set - analyzed_set
    if missing:
        print(f"\n  ATTENZIONE: {len(missing)} URL non trovati nei risultati!")
        if len(missing) <= 10:
            for m in missing:
                print(f"    - {m}")
        else:
            print(f"    (primi 10:)")
            for m in list(missing)[:10]:
                print(f"    - {m}")
    print()

    # Rileva duplicati
    print("[4/5] Rilevamento duplicati...")
    print("-" * 70)

    # Costruisci mappa hash -> lista URL
    hash_to_urls = {}
    for url, h in all_results.items():
        if h not in hash_to_urls:
            hash_to_urls[h] = []
        hash_to_urls[h].append(url)

    # Trova duplicati
    duplicates = []
    duplicate_urls_to_remove = set()

    for h, url_list in hash_to_urls.items():
        if len(url_list) > 1:
            # Il primo nell'ordine originale e' l'originale
            # Ordina per posizione nell'Excel originale
            url_list_sorted = sorted(url_list, key=lambda u: original_urls.index(u)
                                     if u in original_urls else float('inf'))
            original = url_list_sorted[0]
            for dup in url_list_sorted[1:]:
                duplicates.append({
                    "hash": h,
                    "original_url": original,
                    "duplicate_url": dup
                })
                duplicate_urls_to_remove.add(dup)

    print(f"  Hash unici calcolati: {len(hash_to_urls)}")
    print(f"  Gruppi con duplicati: {sum(1 for v in hash_to_urls.values() if len(v) > 1)}")
    print(f"  URL duplicati da rimuovere: {len(duplicate_urls_to_remove)}")
    print()

    # Salva output
    print("[5/5] Creazione file di output...")
    print("-" * 70)

    # 1. JSON duplicati
    with open(DUPLICATES_FILE, "w", encoding="utf-8") as f:
        json.dump(duplicates, f, indent=2, ensure_ascii=False)
    print(f"  JSON duplicati: {DUPLICATES_FILE}")
    print(f"    -> {len(duplicates)} coppie di duplicati")
    print()

    # 2. Excel pulito (senza duplicati, mantieni ordine originale)
    clean_urls = [url for url in original_urls if url not in duplicate_urls_to_remove]
    # Includi anche URL con errore (non sappiamo se sono duplicati, li lasciamo)
    save_clean_excel(clean_urls, OUTPUT_CLEAN_EXCEL)
    print(f"  Excel pulito: {OUTPUT_CLEAN_EXCEL}")
    print(f"    -> {len(clean_urls)} server (rimossi {len(duplicate_urls_to_remove)} duplicati)")
    print()

    # 3. Excel hash-URL mapping (per analisi incrementale futura)
    # Include TUTTI i repo con hash valido
    hash_url_pairs = [(h, url) for url, h in all_results.items()]
    hash_url_pairs.sort(key=lambda x: x[0])  # Ordina per hash
    save_hash_mapping_excel(hash_url_pairs, OUTPUT_HASH_EXCEL)
    print(f"  Excel hash-URL: {OUTPUT_HASH_EXCEL}")
    print(f"    -> {len(hash_url_pairs)} mappature hash-URL")
    print()

    # Riepilogo finale
    print("=" * 70)
    print("  RIEPILOGO FINALE")
    print("=" * 70)
    print(f"  Server originali:              {len(original_urls)}")
    print(f"  Server con hash calcolato:     {len(all_results)}")
    print(f"  Server con errore:             {len(errors)}")
    print(f"  Server non analizzati:         {len(missing)}")
    print(f"  Hash unici:                    {len(hash_to_urls)}")
    print(f"  Duplicati rimossi:             {len(duplicate_urls_to_remove)}")
    print(f"  Server nel file pulito:        {len(clean_urls)}")
    print()
    print("  File creati:")
    print(f"    - {DUPLICATES_FILE}")
    print(f"    - {OUTPUT_CLEAN_EXCEL}")
    print(f"    - {OUTPUT_HASH_EXCEL}")
    print()
    print("  NOTA: hash_url_mapping.xlsx serve per analisi incrementali future.")
    print("        Quando aggiungi nuovi server, calcola solo i loro hash e")
    print("        confrontali con quelli gia' presenti in questo file.")
    print("=" * 70)


if __name__ == "__main__":
    main()
