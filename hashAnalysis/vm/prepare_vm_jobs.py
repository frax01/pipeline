"""
STEP 1: Preparazione dei job per le 9 VM.

Questo script:
  1. Legge l'Excel con i 66333 server
  2. Divide gli URL in 9 parti uguali
  3. Crea 9 file Excel (vm_1.xlsx, vm_2.xlsx, ..., vm_9.xlsx) ognuno con la sua porzione
  4. Crea anche la cartella vm_results/ dove andranno messi i risultati delle VM

Ogni file Excel contiene gli URL nella colonna A.
Questi file vanno copiati sulle rispettive VM insieme a vm_worker.py.

Dopo l'esecuzione su tutte le VM, copiare i file risultanti (vm_X_results.xlsx)
nella cartella vm_results/ e lanciare merge_vm_results.py.
"""

import os
import math
import openpyxl
import openpyxl.styles

# === CONFIGURAZIONE ===
SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
INPUT_FILE = os.path.join(SCRIPT_DIR, "0.0. All servers without duplicates, .git and _ (66333).xlsx")
JOBS_DIR = os.path.join(SCRIPT_DIR, "vm_jobs")
RESULTS_DIR = os.path.join(SCRIPT_DIR, "vm_results")
NUM_VMS = 9


def read_urls_from_excel(filepath):
    """Legge tutti gli URL dalla prima colonna del primo foglio."""
    wb = openpyxl.load_workbook(filepath, read_only=True, data_only=True)
    ws = wb.active
    urls = []
    for row in ws.iter_rows(min_row=2, min_col=1, max_col=1, values_only=True):
        val = row[0]
        if val is not None:
            url = str(val).strip()
            if url:
                urls.append(url)
    wb.close()
    return urls


def save_job_excel(urls, filepath, vm_id):
    """Salva un Excel con gli URL assegnati a una VM."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f"VM {vm_id} - URLs"
    header_font = openpyxl.styles.Font(bold=True, size=12)

    ws["A1"] = "GitHub URL"
    ws["A1"].font = header_font
    ws.column_dimensions["A"].width = 80

    for i, url in enumerate(urls, start=2):
        ws.cell(row=i, column=1, value=url)

    wb.save(filepath)


def main():
    print("=" * 70)
    print("  PREPARE VM JOBS - Divisione server per 9 VM")
    print("=" * 70)
    print()

    # Leggi URL
    print("[1/3] Lettura URL dal file Excel...")
    urls = read_urls_from_excel(INPUT_FILE)
    total = len(urls)
    print(f"  URL totali: {total}")
    print()

    # Crea cartelle
    os.makedirs(JOBS_DIR, exist_ok=True)
    os.makedirs(RESULTS_DIR, exist_ok=True)
    print(f"  Cartella job: {JOBS_DIR}")
    print(f"  Cartella risultati (per dopo): {RESULTS_DIR}")
    print()

    # Dividi in 9 parti
    print(f"[2/3] Divisione in {NUM_VMS} parti...")
    chunk_size = math.ceil(total / NUM_VMS)

    for vm_id in range(1, NUM_VMS + 1):
        start = (vm_id - 1) * chunk_size
        end = min(start + chunk_size, total)
        chunk = urls[start:end]

        filename = f"vm_{vm_id}.xlsx"
        filepath = os.path.join(JOBS_DIR, filename)
        save_job_excel(chunk, filepath, vm_id)

        print(f"  VM {vm_id}: {len(chunk)} URL (indice {start}-{end-1}) -> {filename}")

    # Riepilogo
    print()
    print(f"[3/3] Riepilogo")
    print("-" * 70)
    print(f"  Server totali: {total}")
    print(f"  VM: {NUM_VMS}")
    print(f"  Server per VM: ~{chunk_size}")
    print()
    print("  ISTRUZIONI:")
    print("  1. Copia su ogni VM il file vm_worker.py + il rispettivo vm_X.xlsx")
    print("     Ad esempio sulla VM 1: vm_worker.py + vm_1.xlsx")
    print("  2. Su ogni VM lancia:")
    print("       python vm_worker.py vm_X.xlsx")
    print("     (dove X e' il numero della VM)")
    print("  3. Quando tutte le VM hanno finito, copia i file vm_X_results.xlsx")
    print(f"     nella cartella: {RESULTS_DIR}")
    print("  4. Lancia merge_vm_results.py per unire tutto e trovare i duplicati")
    print("=" * 70)


if __name__ == "__main__":
    main()
