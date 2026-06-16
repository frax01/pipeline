"""
STEP 2: Worker script da eseguire su ogni VM.

Questo script:
  1. Legge un file Excel con gli URL assegnati (passato come argomento)
  2. Per ogni URL: clona il repo, calcola hash SHA-256, elimina il clone
  3. Salva un Excel risultante con colonna A = URL, colonna B = hash
  4. I repo con errore vengono segnati con hash "ERROR" + motivo

Uso:
  python vm_worker.py vm_1.xlsx
  python vm_worker.py vm_2.xlsx
  ...

Il file di output sara' vm_X_results.xlsx (stesso nome ma con _results).

Caratteristiche:
  - Riprendibile: salva progresso in vm_X_progress.json
  - Mostra progresso con percentuale e tempo stimato
  - Pulizia immediata del clone dopo ogni hash
  - Parallelismo: usa N worker thread per velocizzare il processo
"""

import os
import sys
import json
import hashlib
import shutil
import subprocess
import time
import openpyxl
import openpyxl.styles
from concurrent.futures import ThreadPoolExecutor, as_completed
import threading

# === CONFIGURAZIONE ===
SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))

# Numero di worker paralleli per VM
# Consigliato: 4-8 a seconda delle risorse della VM
NUM_WORKERS = 6

# Cartelle e file da escludere dal calcolo hash
EXCLUDED_DIRS = {".git", "node_modules", "dist", "__pycache__", ".venv", "venv",
                 ".env", ".tox", ".mypy_cache", ".pytest_cache", "build", "egg-info"}
EXCLUDED_FILES = {".DS_Store", "Thumbs.db", "desktop.ini"}

# Timeout per il clone (secondi)
CLONE_TIMEOUT = 120
CLONE_DEPTH = 1

# Lock per la scrittura del progresso
progress_lock = threading.Lock()


def compute_server_hash(root, algo="sha256"):
    """
    Calcola un hash deterministico del contenuto di un repository.
    Esclude .git, node_modules, dist, __pycache__, .venv, .DS_Store.
    Ordina i file per path relativo per garantire determinismo.
    """
    h = hashlib.new(algo)
    file_entries = []

    for dirpath, dirnames, filenames in os.walk(root):
        dirnames[:] = [d for d in dirnames if d not in EXCLUDED_DIRS
                       and not d.endswith(".egg-info")]

        for fname in filenames:
            if fname in EXCLUDED_FILES:
                continue

            full_path = os.path.join(dirpath, fname)
            rel_path = os.path.relpath(full_path, root).replace("\\", "/")

            try:
                with open(full_path, "rb") as f:
                    content = f.read()
                file_entries.append((rel_path, content))
            except (PermissionError, OSError):
                continue

    file_entries.sort(key=lambda x: x[0])

    for rel_path, content in file_entries:
        h.update(rel_path.encode("utf-8"))
        h.update(content)

    return h.hexdigest()


def clone_repo(url, dest):
    """Clona un repository con shallow clone (depth=1)."""
    try:
        # GIT_TERMINAL_PROMPT=0 evita che git chieda username/password
        # per repo privati/cancellati, facendolo fallire subito
        env = os.environ.copy()
        env["GIT_TERMINAL_PROMPT"] = "0"
        result = subprocess.run(
            ["git", "clone", "--depth", str(CLONE_DEPTH), "--quiet", url, dest],
            capture_output=True,
            text=True,
            timeout=CLONE_TIMEOUT,
            env=env
        )
        return result.returncode == 0, ""
    except subprocess.TimeoutExpired:
        return False, "timeout"
    except Exception as e:
        return False, str(e)


def cleanup_clone(dest):
    """Elimina la cartella del clone."""
    try:
        if os.path.exists(dest):
            def remove_readonly(func, path, _):
                os.chmod(path, 0o777)
                func(path)
            shutil.rmtree(dest, onerror=remove_readonly)
    except Exception:
        pass


def process_single_repo(url, task_id):
    """
    Processa un singolo repo: clone -> hash -> cleanup.
    Ritorna (url, hash) oppure (url, "ERROR:motivo").
    Ogni task ha la sua cartella unica (basata su task_id).
    """
    clone_dir = os.path.join(SCRIPT_DIR, f"_temp_clone_{task_id}")

    # Pulizia preventiva (caso di interruzione precedente)
    cleanup_clone(clone_dir)

    # Clone
    success, error_msg = clone_repo(url, clone_dir)
    if not success:
        cleanup_clone(clone_dir)
        return (url, f"ERROR:clone_failed:{error_msg}")

    # Hash
    try:
        repo_hash = compute_server_hash(clone_dir)
    except Exception as e:
        cleanup_clone(clone_dir)
        return (url, f"ERROR:hash_failed:{e}")

    # Pulizia immediata
    cleanup_clone(clone_dir)

    return (url, repo_hash)


def format_time(seconds):
    """Formatta secondi in ore:minuti:secondi."""
    if seconds < 0:
        return "??:??:??"
    h = int(seconds // 3600)
    m = int((seconds % 3600) // 60)
    s = int(seconds % 60)
    if h > 0:
        return f"{h}h {m:02d}m {s:02d}s"
    elif m > 0:
        return f"{m}m {s:02d}s"
    else:
        return f"{s}s"


def read_urls_from_excel(filepath):
    """Legge tutti gli URL dalla prima colonna."""
    wb = openpyxl.load_workbook(filepath, read_only=True, data_only=True)
    ws = wb.active
    urls = []
    for row in ws.iter_rows(min_row=2, min_col=1, max_col=1, values_only=True):
        val = row[0]
        if val is not None:
            url = str(val).strip()
            if url:
                urls.append(url)
    wb.close()
    return urls


def load_progress(progress_file):
    """Carica il progresso precedente."""
    if os.path.exists(progress_file):
        try:
            with open(progress_file, "r", encoding="utf-8") as f:
                data = json.load(f)
            return data  # {url: hash_or_error, ...}
        except (json.JSONDecodeError, KeyError):
            pass
    return {}


def save_progress(progress_file, completed):
    """Salva il progresso su disco (thread-safe)."""
    with progress_lock:
        with open(progress_file, "w", encoding="utf-8") as f:
            json.dump(completed, f, ensure_ascii=False)


def save_results_excel(completed, filepath):
    """Salva Excel con URL nella colonna A e hash nella colonna B."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Hash Results"
    header_font = openpyxl.styles.Font(bold=True, size=12)

    ws["A1"] = "GitHub URL"
    ws["A1"].font = header_font
    ws["B1"] = "SHA-256 Hash"
    ws["B1"].font = header_font
    ws.column_dimensions["A"].width = 80
    ws.column_dimensions["B"].width = 70

    row = 2
    for url, hash_val in sorted(completed.items(), key=lambda x: x[0].lower()):
        ws.cell(row=row, column=1, value=url)
        ws.cell(row=row, column=2, value=hash_val)
        row += 1

    wb.save(filepath)


def main():
    # Verifica argomento
    if len(sys.argv) < 2:
        print("Uso: python vm_worker.py <file_excel_input>")
        print("Esempio: python vm_worker.py vm_1.xlsx")
        sys.exit(1)

    input_filename = sys.argv[1]
    input_file = os.path.join(SCRIPT_DIR, input_filename)

    if not os.path.exists(input_file):
        print(f"ERRORE: File non trovato: {input_file}")
        sys.exit(1)

    # Ricava il nome base per i file di output
    base_name = os.path.splitext(input_filename)[0]  # es. "vm_1"
    progress_file = os.path.join(SCRIPT_DIR, f"{base_name}_progress.json")
    output_file = os.path.join(SCRIPT_DIR, f"{base_name}_results.xlsx")

    print("=" * 70)
    print(f"  VM WORKER - Analisi hash repository")
    print(f"  Input: {input_filename}")
    print(f"  Worker paralleli: {NUM_WORKERS}")
    print("=" * 70)
    print()

    # Leggi URL
    print("[1/3] Lettura URL...")
    urls = read_urls_from_excel(input_file)
    total = len(urls)
    print(f"  URL da analizzare: {total}")

    # Carica progresso
    print("[2/3] Verifica progresso precedente...")
    completed = load_progress(progress_file)
    already_done = sum(1 for url in urls if url in completed)
    remaining = total - already_done
    print(f"  Gia' completati: {already_done}")
    print(f"  Da analizzare: {remaining}")

    if remaining == 0:
        print("\n  Tutti i repo sono gia' stati analizzati!")
        print("  Salvataggio Excel risultati...")
        save_results_excel(completed, output_file)
        print(f"  File salvato: {output_file}")
        return

    # Filtra solo quelli da fare
    urls_to_process = [url for url in urls if url not in completed]

    print(f"\n[3/3] Analisi hash con {NUM_WORKERS} worker paralleli...")
    print("-" * 70)

    start_time = time.time()
    processed = 0
    errors = 0

    # Conta complessiva per il progresso
    counter_lock = threading.Lock()

    def update_and_print(url, result):
        nonlocal processed, errors, completed
        processed += 1
        is_error = result.startswith("ERROR:")
        if is_error:
            errors += 1

        overall_done = already_done + processed
        pct = (overall_done / total) * 100

        elapsed = time.time() - start_time
        if processed > 0:
            avg_time = elapsed / processed
            eta = avg_time * (remaining - processed)
            eta_str = format_time(eta)
        else:
            eta_str = "calcolo..."

        short_url = url if len(url) <= 45 else url[:42] + "..."
        status = "ERRORE" if is_error else result[:12] + "..."

        print(f"  [{overall_done}/{total}] ({pct:.1f}%) ETA: {eta_str} | {short_url} -> {status}")

        # Aggiorna progresso
        with progress_lock:
            completed[url] = result

        # Salva progresso ogni 10 repo
        if processed % 10 == 0:
            save_progress(progress_file, completed)

    # Esecuzione parallela
    with ThreadPoolExecutor(max_workers=NUM_WORKERS) as executor:
        # Sottometti tutti i job
        future_to_url = {}
        for i, url in enumerate(urls_to_process):
            future = executor.submit(process_single_repo, url, i)
            future_to_url[future] = url

        # Raccogli risultati man mano che arrivano
        for future in as_completed(future_to_url):
            url = future_to_url[future]
            try:
                result_url, result_hash = future.result()
                with counter_lock:
                    update_and_print(result_url, result_hash)
            except Exception as e:
                with counter_lock:
                    update_and_print(url, f"ERROR:exception:{e}")

    # Salvataggio finale del progresso
    save_progress(progress_file, completed)

    print("-" * 70)
    elapsed_total = time.time() - start_time
    print(f"\n  Analisi completata in {format_time(elapsed_total)}")
    print(f"  Repo processati: {processed}")
    print(f"  Errori: {errors}")
    print(f"  Velocita' media: {elapsed_total/max(processed,1):.1f}s per repo")

    # Salva Excel risultati
    print(f"\n  Salvataggio risultati in {output_file}...")
    save_results_excel(completed, output_file)

    print()
    print("=" * 70)
    print(f"  COMPLETATO!")
    print(f"  File risultati: {output_file}")
    print(f"  Repo totali: {total}")
    print(f"  Con hash: {total - errors}")
    print(f"  Con errore: {errors}")
    print()
    print(f"  PROSSIMO PASSO:")
    print(f"  Copia {os.path.basename(output_file)} nella cartella vm_results/")
    print(f"  sul PC principale e lancia merge_vm_results.py")
    print("=" * 70)


if __name__ == "__main__":
    main()
