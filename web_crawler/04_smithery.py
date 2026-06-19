import requests
import time
import os
from openpyxl import Workbook, load_workbook

API_KEY = os.environ.get("SMITHERY_API_KEY", "")

BASE_URL = "https://registry.smithery.ai"

OUTPUT_FILE = "04_smithery.xlsx"

START_PAGE = 75          # <-- CAMBIA QUI se vuoi ripartire (es: 32)
SLEEP_BETWEEN_PAGES = 1 # secondi
SLEEP_BETWEEN_DETAILS = 0.5

MAX_RETRIES = 5

headers = {
    "Authorization": f"Bearer {API_KEY}",
    "Accept": "application/json"
}

# =========================
# FUNZIONI API
# =========================

def fetch_page(page):
    for attempt in range(MAX_RETRIES):
        r = requests.get(
            f"{BASE_URL}/servers",
            headers=headers,
            params={"page": page}
        )

        if r.status_code == 429:
            wait = 2 ** attempt
            print(f"[429] Page {page} – attendo {wait}s")
            time.sleep(wait)
            continue

        r.raise_for_status()
        return r.json()

    raise RuntimeError(f"Rate limit persistente su page {page}")

def fetch_details(qname):
    for attempt in range(MAX_RETRIES):
        r = requests.get(
            f"{BASE_URL}/servers/{qname}",
            headers=headers
        )

        if r.status_code == 429:
            wait = 2 ** attempt
            print(f"[429] Details {qname} – attendo {wait}s")
            time.sleep(wait)
            continue

        r.raise_for_status()
        return r.json()

    raise RuntimeError(f"Rate limit persistente su details {qname}")

def extract_github_url(details):
    if isinstance(details.get("repository"), str):
        if "github.com" in details["repository"]:
            return details["repository"]

    source = details.get("source", {})
    if isinstance(source, dict):
        url = source.get("url")
        if isinstance(url, str) and "github.com" in url:
            return url

    return ""

# =========================
# EXCEL: CREA O RIPRENDI
# =========================

if os.path.exists(OUTPUT_FILE):
    wb = load_workbook(OUTPUT_FILE)
    ws = wb.active
    print("[+] Excel esistente trovato, riprendo")

    existing = {
        row[0] for row in ws.iter_rows(min_row=2, values_only=True)
        if row[0]
    }
else:
    wb = Workbook()
    ws = wb.active
    ws.title = "Smithery Servers"

    ws.append([
        "qualifiedName",
        "displayName",
        "smithery_url",
        "github_repo"
    ])

    existing = set()
    print("[+] Nuovo Excel creato")

# =========================
# RACCOLTA
# =========================

print(f"[+] Parto da pagina {START_PAGE}")

try:
    first_page = fetch_page(START_PAGE)
    total_pages = first_page["pagination"]["totalPages"]
except Exception as e:
    print(f"[!] Errore iniziale: {e}")
    raise

print(f"[+] Pagine totali: {total_pages}")

for page in range(START_PAGE, total_pages + 1):
    print(f"[+] Pagina {page}/{total_pages}")

    try:
        data = fetch_page(page)
    except Exception as e:
        print(f"[!] Interrotto a pagina {page}: {e}")
        break

    for server in data["servers"]:
        qname = server["qualifiedName"]

        if qname in existing:
            continue

        smithery_url = server.get(
            "homepage",
            f"https://smithery.ai/server/{qname}"
        )

        try:
            details = fetch_details(qname)
            github_url = extract_github_url(details)
        except Exception as e:
            print(f"[!] Skip dettagli {qname}: {e}")
            github_url = ""

        ws.append([
            qname,
            server.get("displayName", ""),
            smithery_url,
            github_url
        ])

        existing.add(qname)

        # SALVA SUBITO (fondamentale)
        wb.save(OUTPUT_FILE)

        time.sleep(SLEEP_BETWEEN_DETAILS)

    time.sleep(SLEEP_BETWEEN_PAGES)

print(f"Fine. Excel salvato in: {OUTPUT_FILE}")