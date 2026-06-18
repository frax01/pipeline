"""
Scrape all MCP server GitHub links from glama.ai/mcp/servers
and save them to an Excel file.

The script:
1. Fetches the sitemap XML to get all server page URLs
2. Extracts @author/repo patterns from URLs
3. Constructs GitHub repository links
4. Saves all links to an Excel file (first column)
"""

import subprocess
import xml.etree.ElementTree as ET
import re
import os
import time
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill

SITEMAP_URL = "https://glama.ai/sitemaps/mcp-servers.xml"
OUTPUT_DIR = os.path.dirname(os.path.abspath(__file__))
OUTPUT_FILE = os.path.join(OUTPUT_DIR, "06_glama.xlsx")

# Pattern to extract @author/repo from glama.ai URLs
PATTERN = re.compile(r"/mcp/servers/@([^/]+)/([^/]+)")


def fetch_sitemap(url):
    """Fetch the sitemap XML using curl (handles HTTP 103 Early Hints)."""
    print(f"[*] Fetching sitemap from {url} ...")
    result = subprocess.run(
        ["curl", "-s", "-L", url, "-H", "User-Agent: Mozilla/5.0"],
        capture_output=True, text=True, timeout=120
    )
    if result.returncode != 0:
        raise RuntimeError(f"curl failed with code {result.returncode}: {result.stderr}")
    xml_text = result.stdout
    print(f"[+] Sitemap fetched successfully ({len(xml_text):,} bytes)")
    return xml_text


def parse_server_urls(xml_text):
    """Parse <loc> elements from the sitemap XML and extract GitHub links."""
    # Remove XML namespace to simplify parsing
    xml_text = re.sub(r'\sxmlns="[^"]+"', '', xml_text, count=1)

    root = ET.fromstring(xml_text)
    github_links = []
    seen = set()

    for url_elem in root.findall(".//url/loc"):
        loc = url_elem.text
        if not loc:
            continue
        match = PATTERN.search(loc)
        if match:
            author = match.group(1)
            repo = match.group(2)
            github_url = f"https://github.com/{author}/{repo}"
            if github_url not in seen:
                seen.add(github_url)
                github_links.append(github_url)

    return github_links


def save_to_excel(github_links, output_path):
    """Save the GitHub links to an Excel file."""
    wb = Workbook()
    ws = wb.active
    ws.title = "MCP Servers GitHub Links"

    # Header
    header_cell = ws.cell(row=1, column=1, value="GitHub Repository URL")
    header_cell.font = Font(bold=True, size=12)
    header_cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_cell.font = Font(bold=True, size=12, color="FFFFFF")
    header_cell.alignment = Alignment(horizontal="center")

    # Add count in column B header
    count_cell = ws.cell(row=1, column=2, value=f"Total: {len(github_links)}")
    count_cell.font = Font(bold=True, size=11)

    # Data
    for i, link in enumerate(github_links, start=2):
        ws.cell(row=i, column=1, value=link)

    # Adjust column width
    ws.column_dimensions["A"].width = 80
    ws.column_dimensions["B"].width = 18

    wb.save(output_path)
    print(f"[+] Saved {len(github_links)} GitHub links to: {output_path}")


def main():
    start = time.time()

    # Step 1: Fetch sitemap
    xml_text = fetch_sitemap(SITEMAP_URL)

    # Step 2: Parse and extract GitHub links
    github_links = parse_server_urls(xml_text)
    print(f"[+] Extracted {len(github_links)} unique GitHub repository links")

    if not github_links:
        print("[!] No links found. The sitemap structure may have changed.")
        return

    # Show a few examples
    print("\n[*] Sample links:")
    for link in github_links[:10]:
        print(f"    {link}")
    if len(github_links) > 10:
        print(f"    ... and {len(github_links) - 10} more")

    # Step 3: Save to Excel
    save_to_excel(github_links, OUTPUT_FILE)

    elapsed = time.time() - start
    print(f"\n[+] Done in {elapsed:.1f}s")


if __name__ == "__main__":
    main()
