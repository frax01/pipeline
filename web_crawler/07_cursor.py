"""
Scraper per estrarre tutti i link GitHub degli MCP server da cursor.directory/mcp
Usa Playwright per il rendering JavaScript e lo scroll infinito.
Salva i risultati in un file Excel.
"""

import re
import os
import time
import openpyxl
import openpyxl.styles
from playwright.sync_api import sync_playwright

BASE_URL = "https://cursor.directory"
MCP_LIST_URL = f"{BASE_URL}/mcp"
OUTPUT_DIR = os.path.dirname(os.path.abspath(__file__))
OUTPUT_FILE = os.path.join(OUTPUT_DIR, "07_cursor.xlsx")

GH_REPO_PATTERN = re.compile(r'https?://github\.com/[A-Za-z0-9_.\-]+/[A-Za-z0-9_.\-]+')
EXCLUDE_REPOS = {"https://github.com/pontusab/cursor.directory"}


def clean_github_url(url):
    """Normalizza un URL GitHub al repo base."""
    m = GH_REPO_PATTERN.match(url)
    if m:
        clean = m.group(0).rstrip("/.")
        if clean not in EXCLUDE_REPOS:
            return clean
    return None


def scroll_and_collect(page):
    """Scrolla la pagina fino in fondo per caricare tutti gli MCP, poi raccoglie i link."""
    print("  Scrolling per caricare tutti gli MCP server...")

    prev_height = 0
    no_change_count = 0
    scroll_count = 0

    while no_change_count < 5:
        page.evaluate("window.scrollTo(0, document.body.scrollHeight)")
        time.sleep(1.5)
        scroll_count += 1

        current_height = page.evaluate("document.body.scrollHeight")

        # Conta quanti MCP card sono visibili
        mcp_count = page.evaluate("""
            () => {
                const links = document.querySelectorAll('a[href^="/mcp/"]');
                const slugs = new Set();
                links.forEach(a => {
                    const match = a.href.match(/\\/mcp\\/([a-zA-Z0-9_\\-]+)/);
                    if (match && !['new', 'install', 'servers'].includes(match[1])) {
                        slugs.add(match[1]);
                    }
                });
                return slugs.size;
            }
        """)

        print(f"  Scroll #{scroll_count}: altezza={current_height}, MCP visibili={mcp_count}")

        if current_height == prev_height:
            no_change_count += 1
        else:
            no_change_count = 0
        prev_height = current_height

    print(f"  Scroll completato dopo {scroll_count} iterazioni")


def collect_mcp_slugs(page):
    """Raccoglie tutti gli slug MCP dalla pagina."""
    slugs = page.evaluate("""
        () => {
            const links = document.querySelectorAll('a[href^="/mcp/"]');
            const slugs = new Set();
            links.forEach(a => {
                const match = a.href.match(/\\/mcp\\/([a-zA-Z0-9_\\-]+)/);
                if (match && !['new', 'install', 'servers'].includes(match[1])) {
                    slugs.add(match[1]);
                }
            });
            return Array.from(slugs);
        }
    """)
    return slugs


def extract_github_from_page_source(page):
    """Estrae link GitHub dal contenuto HTML renderizzato della pagina."""
    content = page.content()
    github_links = set()

    for m in GH_REPO_PATTERN.finditer(content):
        clean = clean_github_url(m.group(0))
        if clean:
            github_links.add(clean)

    return github_links


def extract_github_from_mcp_page(page, slug):
    """Visita una pagina MCP individuale e cerca il link GitHub."""
    url = f"{BASE_URL}/mcp/{slug}"
    try:
        page.goto(url, wait_until="networkidle", timeout=15000)
        time.sleep(1)
    except Exception as e:
        print(f"    Errore caricamento: {e}")
        return set()

    github_links = set()

    # Cerca link GitHub nel DOM renderizzato
    links = page.evaluate("""
        () => {
            const allLinks = document.querySelectorAll('a[href*="github.com"]');
            return Array.from(allLinks).map(a => a.href);
        }
    """)

    for link in links:
        clean = clean_github_url(link)
        if clean:
            github_links.add(clean)

    # Cerca anche nel testo della pagina (a volte i link sono nel testo ma non <a>)
    content = page.content()
    for m in GH_REPO_PATTERN.finditer(content):
        clean = clean_github_url(m.group(0))
        if clean:
            github_links.add(clean)

    return github_links


def save_to_excel(github_links):
    """Salva i link GitHub ordinati in un file Excel."""
    sorted_links = sorted(github_links, key=str.lower)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "MCP GitHub Links"

    header_font = openpyxl.styles.Font(bold=True, size=12)
    ws["A1"] = "GitHub URL"
    ws["A1"].font = header_font
    ws.column_dimensions["A"].width = 75

    for i, link in enumerate(sorted_links, start=2):
        ws.cell(row=i, column=1, value=link)

    wb.save(OUTPUT_FILE)
    print(f"  File salvato: {OUTPUT_FILE}")


def main():
    print("=" * 60)
    print("  CURSOR DIRECTORY - MCP GitHub Links Scraper")
    print("=" * 60)
    print()

    all_github_links = set()

    with sync_playwright() as p:
        browser = p.chromium.launch(headless=True)
        context = browser.new_context(
            user_agent="Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 "
                       "(KHTML, like Gecko) Chrome/131.0.0.0 Safari/537.36",
            viewport={"width": 1920, "height": 1080},
        )
        page = context.new_page()

        # Step 1: Carica pagina principale e scrolla per tutti gli MCP
        print("[1/3] Caricamento pagina principale...")
        page.goto(MCP_LIST_URL, wait_until="networkidle", timeout=30000)
        time.sleep(2)

        # Raccogli link GitHub dalla pagina principale (prima dello scroll)
        gh_before = extract_github_from_page_source(page)
        all_github_links.update(gh_before)
        print(f"  GitHub trovati prima dello scroll: {len(gh_before)}")

        # Scrolla per caricare tutti i contenuti
        scroll_and_collect(page)

        # Raccogli dopo lo scroll
        gh_after = extract_github_from_page_source(page)
        all_github_links.update(gh_after)
        print(f"  GitHub trovati dopo scroll: {len(gh_after)}")

        # Raccogli tutti gli slug
        slugs = collect_mcp_slugs(page)
        print(f"  Slug MCP totali: {len(slugs)}")

        # Step 2: Visita ogni pagina MCP individuale
        print(f"\n[2/3] Scansione di {len(slugs)} pagine MCP individuali...")
        for i, slug in enumerate(sorted(slugs), 1):
            print(f"  [{i}/{len(slugs)}] {slug}...", end=" ", flush=True)
            page_links = extract_github_from_mcp_page(page, slug)
            new = page_links - all_github_links
            all_github_links.update(page_links)
            if page_links:
                print(f"{len(page_links)} link (nuovi: {len(new)})")
            else:
                print("nessun link GitHub")
            time.sleep(0.5)

        browser.close()

    # Step 3: Salva in Excel
    print(f"\n[3/3] Salvataggio di {len(all_github_links)} link GitHub...")
    if all_github_links:
        save_to_excel(all_github_links)
    else:
        print("  ERRORE: Nessun link GitHub trovato!")
        return

    print(f"\n{'=' * 60}")
    print(f"  COMPLETATO! {len(all_github_links)} link GitHub salvati")
    print(f"  File: {OUTPUT_FILE}")
    print(f"{'=' * 60}")


if __name__ == "__main__":
    main()
