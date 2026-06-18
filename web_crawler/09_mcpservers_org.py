"""
Scraper per mcpservers.org - Estrae tutti i link GitHub dei server MCP
e li salva in un file Excel.
"""

import requests
from bs4 import BeautifulSoup
import openpyxl
import time
import re
import sys


BASE_URL = "https://mcpservers.org"
ALL_SERVERS_URL = f"{BASE_URL}/all"
HEADERS = {
    "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 "
                  "(KHTML, like Gecko) Chrome/131.0.0.0 Safari/537.36"
}
OUTPUT_FILE = "09_mcpservers_org.xlsx"


def get_total_pages(session):
    """Get the total number of pages from the first page."""
    resp = session.get(ALL_SERVERS_URL, headers=HEADERS, timeout=30)
    resp.raise_for_status()
    soup = BeautifulSoup(resp.text, "html.parser")

    # Look for pagination info - "Page X of Y" or last page link
    # Try to find total from text like "6,231 servers" or pagination links
    pagination = soup.select("nav a, .pagination a, a[href*='page=']")
    max_page = 1
    for link in pagination:
        href = link.get("href", "")
        match = re.search(r'page=(\d+)', href)
        if match:
            page_num = int(match.group(1))
            if page_num > max_page:
                max_page = page_num

    # Also check text content for total
    text = soup.get_text()
    total_match = re.search(r'of\s+([\d,]+)\s+(?:page|total)', text, re.IGNORECASE)
    if total_match:
        pass  # max_page from links is more reliable

    return max_page, soup


def extract_server_links_from_page(soup):
    """Extract all /servers/ links from a page."""
    server_paths = set()
    for link in soup.find_all("a", href=True):
        href = link["href"]
        if href.startswith("/servers/"):
            server_paths.add(href)
    return server_paths


def extract_github_links_from_page(soup):
    """Extract GitHub links directly visible on the page."""
    github_links = set()
    for link in soup.find_all("a", href=True):
        href = link["href"]
        if "github.com" in href and "/servers/" not in href:
            # Clean the URL
            clean = href.split("?")[0].split("#")[0].rstrip("/")
            if clean:
                github_links.add(clean)
    return github_links


def get_github_from_server_page(session, server_path):
    """Visit a server detail page and extract the GitHub link."""
    url = f"{BASE_URL}{server_path}"
    try:
        resp = session.get(url, headers=HEADERS, timeout=20)
        resp.raise_for_status()
        soup = BeautifulSoup(resp.text, "html.parser")
        for link in soup.find_all("a", href=True):
            href = link["href"]
            if "github.com" in href:
                clean = href.split("?")[0].split("#")[0].rstrip("/")
                # Only keep repo-level links (github.com/owner/repo)
                parts = clean.replace("https://github.com/", "").replace("http://github.com/", "").split("/")
                if len(parts) >= 2:
                    return clean
        return None
    except Exception as e:
        print(f"  [ERRORE] {server_path}: {e}")
        return None


def scrape_all_servers():
    """Main scraping function."""
    session = requests.Session()

    print("=" * 60)
    print("MCP Servers GitHub Link Scraper")
    print("Sito: mcpservers.org")
    print("=" * 60)

    # Step 1: Get total pages
    print("\n[1/3] Recupero numero totale di pagine...")
    max_page, first_soup = get_total_pages(session)
    print(f"      Trovate {max_page} pagine da analizzare")

    # Step 2: Collect all server paths from all pages
    print(f"\n[2/3] Raccolta link dei server da tutte le {max_page} pagine...")
    all_server_paths = set()
    all_github_links = set()

    # Process first page
    paths = extract_server_links_from_page(first_soup)
    github_direct = extract_github_links_from_page(first_soup)
    all_server_paths.update(paths)
    all_github_links.update(github_direct)
    print(f"      Pagina 1/{max_page}: {len(paths)} server trovati, {len(github_direct)} link GitHub diretti")

    # Process remaining pages
    for page in range(2, max_page + 1):
        try:
            url = f"{ALL_SERVERS_URL}?page={page}"
            resp = session.get(url, headers=HEADERS, timeout=30)
            resp.raise_for_status()
            soup = BeautifulSoup(resp.text, "html.parser")

            paths = extract_server_links_from_page(soup)
            github_direct = extract_github_links_from_page(soup)
            all_server_paths.update(paths)
            all_github_links.update(github_direct)

            if page % 10 == 0 or page == max_page:
                print(f"      Pagina {page}/{max_page}: totale {len(all_server_paths)} server, {len(all_github_links)} link GitHub")

            time.sleep(0.3)  # Rate limiting
        except Exception as e:
            print(f"      [ERRORE] Pagina {page}: {e}")
            time.sleep(2)

    print(f"\n      Totale server trovati: {len(all_server_paths)}")
    print(f"      Link GitHub trovati direttamente dalle pagine lista: {len(all_github_links)}")

    # Step 3: For server paths that we don't have GitHub links for,
    # try to derive GitHub URL from the path pattern
    # Pattern: /servers/{owner}/{repo} -> https://github.com/{owner}/{repo}
    print(f"\n[3/3] Derivazione link GitHub dai percorsi server...")
    derived_count = 0
    for path in all_server_paths:
        # /servers/owner/repo -> github.com/owner/repo
        parts = path.replace("/servers/", "").split("/")
        if len(parts) >= 2:
            owner = parts[0]
            repo = parts[1]
            # Skip paths that look like they're not GitHub repos
            if "github-com-" in owner or "github-com-" in path:
                # These have a different format, e.g. /servers/github-com-owner-repo
                # Try to parse: github-com-{owner}-{rest}
                combined = path.replace("/servers/", "")
                if combined.startswith("github-com-"):
                    remainder = combined[len("github-com-"):]
                    # This is tricky - we can't reliably split owner/repo
                    # Just skip these, they should be caught by direct links
                    continue
            github_url = f"https://github.com/{owner}/{repo}"
            all_github_links.add(github_url)
            derived_count += 1

    print(f"      Link GitHub derivati dai percorsi: {derived_count}")
    print(f"      Totale link GitHub unici: {len(all_github_links)}")

    # Sort links
    sorted_links = sorted(all_github_links, key=str.lower)

    # Save to Excel
    print(f"\nSalvataggio in {OUTPUT_FILE}...")
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "MCP GitHub Links"

    # Header
    ws["A1"] = "GitHub URL"
    ws["A1"].font = openpyxl.styles.Font(bold=True)
    ws.column_dimensions["A"].width = 80

    # Data
    for i, link in enumerate(sorted_links, start=2):
        ws[f"A{i}"] = link

    wb.save(OUTPUT_FILE)
    print(f"Salvati {len(sorted_links)} link GitHub in '{OUTPUT_FILE}'")
    print("=" * 60)
    print("COMPLETATO!")
    print("=" * 60)

    return sorted_links


if __name__ == "__main__":
    links = scrape_all_servers()
