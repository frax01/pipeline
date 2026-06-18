"""
Script supplementare: visita le pagine dei server con pattern 'github-com-'
per recuperare i link GitHub reali mancanti dal primo scraping.
Aggiorna il file Excel esistente.
"""

import requests
from bs4 import BeautifulSoup
import openpyxl
import time
import re

BASE_URL = "https://mcpservers.org"
ALL_SERVERS_URL = f"{BASE_URL}/all"
HEADERS = {
    "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 "
                  "(KHTML, like Gecko) Chrome/131.0.0.0 Safari/537.36"
}
OUTPUT_FILE = "09_mcpservers_org.xlsx"


def collect_github_com_paths(session):
    """Collect all server paths with the github-com- pattern from all pages."""
    print("Raccolta percorsi 'github-com-' da tutte le pagine...")

    # Get total pages
    resp = session.get(ALL_SERVERS_URL, headers=HEADERS, timeout=30)
    resp.raise_for_status()
    soup = BeautifulSoup(resp.text, "html.parser")

    pagination = soup.select("a[href*='page=']")
    max_page = 1
    for link in pagination:
        href = link.get("href", "")
        match = re.search(r'page=(\d+)', href)
        if match:
            page_num = int(match.group(1))
            if page_num > max_page:
                max_page = page_num

    github_com_paths = set()

    # Process all pages
    for page in range(1, max_page + 1):
        try:
            url = ALL_SERVERS_URL if page == 1 else f"{ALL_SERVERS_URL}?page={page}"
            if page > 1:
                resp = session.get(url, headers=HEADERS, timeout=30)
                resp.raise_for_status()
                soup = BeautifulSoup(resp.text, "html.parser")

            for link in soup.find_all("a", href=True):
                href = link["href"]
                if href.startswith("/servers/") and "github-com-" in href:
                    github_com_paths.add(href)

            if page % 20 == 0 or page == max_page:
                print(f"  Pagina {page}/{max_page}: {len(github_com_paths)} percorsi github-com- trovati")

            time.sleep(0.3)
        except Exception as e:
            print(f"  [ERRORE] Pagina {page}: {e}")
            time.sleep(2)

    return github_com_paths


def get_github_link_from_detail(session, server_path):
    """Visit a server detail page and find the actual GitHub link."""
    url = f"{BASE_URL}{server_path}"
    try:
        resp = session.get(url, headers=HEADERS, timeout=20)
        resp.raise_for_status()
        soup = BeautifulSoup(resp.text, "html.parser")

        for link in soup.find_all("a", href=True):
            href = link["href"]
            if "github.com/" in href and "/servers/" not in href:
                clean = href.split("?")[0].split("#")[0].rstrip("/")
                # Verify it looks like a repo URL
                stripped = clean.replace("https://github.com/", "").replace("http://github.com/", "")
                parts = stripped.split("/")
                if len(parts) >= 2 and parts[0] and parts[1]:
                    return clean
        return None
    except Exception as e:
        return None


def main():
    session = requests.Session()

    print("=" * 60)
    print("Recupero link GitHub mancanti (pattern github-com-)")
    print("=" * 60)

    # Load existing links from Excel
    print(f"\nCaricamento link esistenti da {OUTPUT_FILE}...")
    existing_links = set()
    try:
        wb = openpyxl.load_workbook(OUTPUT_FILE)
        ws = wb.active
        for row in ws.iter_rows(min_row=2, max_col=1, values_only=True):
            if row[0]:
                existing_links.add(row[0])
        print(f"  {len(existing_links)} link esistenti caricati")
    except FileNotFoundError:
        print("  File non trovato, creo nuovo file")
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "MCP GitHub Links"
        ws["A1"] = "GitHub URL"
        ws["A1"].font = openpyxl.styles.Font(bold=True)
        ws.column_dimensions["A"].width = 80

    # Collect github-com- paths
    github_com_paths = collect_github_com_paths(session)
    print(f"\nTrovati {len(github_com_paths)} server con pattern github-com-")

    # Visit each detail page to get the real GitHub link
    print("\nVisita pagine di dettaglio per recuperare link GitHub reali...")
    new_links = set()
    errors = 0

    for i, path in enumerate(sorted(github_com_paths), 1):
        github_url = get_github_link_from_detail(session, path)
        if github_url and github_url not in existing_links:
            new_links.add(github_url)

        if github_url is None:
            errors += 1

        if i % 20 == 0 or i == len(github_com_paths):
            print(f"  Progresso: {i}/{len(github_com_paths)} - Nuovi link: {len(new_links)} - Errori: {errors}")

        time.sleep(0.3)

    print(f"\nNuovi link GitHub trovati: {len(new_links)}")

    if new_links:
        # Merge all links and sort
        all_links = existing_links | new_links
        sorted_links = sorted(all_links, key=str.lower)

        # Rewrite Excel
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "MCP GitHub Links"
        ws["A1"] = "GitHub URL"
        ws["A1"].font = openpyxl.styles.Font(bold=True)
        ws.column_dimensions["A"].width = 80

        for i, link in enumerate(sorted_links, start=2):
            ws[f"A{i}"] = link

        wb.save(OUTPUT_FILE)
        print(f"\nFile aggiornato: {len(sorted_links)} link totali in '{OUTPUT_FILE}'")
    else:
        print("\nNessun nuovo link da aggiungere.")

    print("=" * 60)
    print("COMPLETATO!")
    print("=" * 60)


if __name__ == "__main__":
    main()
