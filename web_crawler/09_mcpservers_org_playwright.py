from playwright.sync_api import sync_playwright
from openpyxl import Workbook, load_workbook
import time
import os

BASE_URL = "https://mcpservers.org"
LIST_URL = "https://mcpservers.org/all?page={}"
OUTPUT_XLSX = "09_mcpservers_org_playwright.xlsx"

TOTAL_PAGES = 190

# -------------------------
# Utils
# -------------------------

def load_or_create_workbook():
    if os.path.exists(OUTPUT_XLSX):
        wb = load_workbook(OUTPUT_XLSX)
        ws = wb.active
        processed = set()
        for row in ws.iter_rows(min_row=2, values_only=True):
            if row and row[1]:
                processed.add(row[1])  # server page url
        print(f"[+] Ripresa: {len(processed)} server già processati")
        return wb, ws, processed
    else:
        wb = Workbook()
        ws = wb.active
        ws.title = "MCP Servers"
        ws.append(["Server name", "Server page", "GitHub repo"])
        return wb, ws, set()

def dismiss_overlay(page):
    page.evaluate("""
    () => {
        document.querySelectorAll(
            '.fc-dialog-overlay, .fc-consent-root'
        ).forEach(el => el.remove());
    }
    """)

# -------------------------
# Main
# -------------------------

def main():
    wb, ws, processed = load_or_create_workbook()

    with sync_playwright() as p:
        browser = p.chromium.launch(headless=True)
        context = browser.new_context()
        page = context.new_page()

        for page_num in range(1, TOTAL_PAGES + 1):
            try:
                list_url = LIST_URL.format(page_num)
                print(f"\n[+] Pagina {page_num}/{TOTAL_PAGES}")

                # 1️⃣ Carica pagina lista
                page.goto(list_url, timeout=60000)
                page.wait_for_load_state("networkidle")
                time.sleep(0.8)
                dismiss_overlay(page)

                # 2️⃣ Estrai href della pagina (PRIMA di navigare altrove)
                cards = page.locator("a[href^='/servers/']")
                count = cards.count()

                page_servers = []
                for i in range(0, count):
                    a = cards.nth(i)
                    href = a.get_attribute("href")
                    name = a.inner_text().split("\n")[0].strip()
                    if href:
                        page_servers.append((name, BASE_URL + href))

                print(f"    → Server trovati: {len(page_servers)}")

                # 3️⃣ Visita i server della pagina
                for name, server_url in page_servers:

                    if server_url in processed:
                        continue

                    print(f"        • {name}")

                    try:
                        page.goto(server_url, timeout=60000)
                        page.wait_for_load_state("networkidle")
                        time.sleep(0.6)
                        dismiss_overlay(page)

                        github_link = "link non trovato"
                        gh = page.locator("a[href*='github.com']")
                        if gh.count() > 0:
                            github_link = gh.first.get_attribute("href")

                        ws.append([name, server_url, github_link])
                        wb.save(OUTPUT_XLSX)
                        processed.add(server_url)

                    except Exception as e:
                        print(f"          ! errore server: {e}")
                        ws.append([name, server_url, "errore"])
                        wb.save(OUTPUT_XLSX)
                        processed.add(server_url)

            except KeyboardInterrupt:
                print("\n[!] Interruzione manuale, stato salvato.")
                wb.save(OUTPUT_XLSX)
                browser.close()
                return

        browser.close()

    wb.save(OUTPUT_XLSX)
    print("\n[✓] Completato")

# -------------------------
# Entry point
# -------------------------

if __name__ == "__main__":
    main()
