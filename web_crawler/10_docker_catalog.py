"""
Script per scaricare tutti i link GitHub degli MCP server
dalla pagina Docker Hub: https://hub.docker.com/mcp/explore

Strategia:
  1. Usa Selenium per visitare tutte le categorie e raccogliere i nomi dei server
     (dal sorgente HTML, pattern /catalog/v3/readme/<nome>.md)
  2. Per ogni server, scarica il readme da desktop.docker.com/mcp/catalog/v3/readme/<nome>.md
  3. Estrae il link Repository dal readme
  4. Salva tutti i link GitHub in un file Excel

Requisiti:
    pip install selenium openpyxl webdriver-manager requests
"""

import time
import re
import sys
import os
import requests
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment

try:
    from selenium import webdriver
    from selenium.webdriver.chrome.options import Options
    from selenium.webdriver.chrome.service import Service
    from selenium.webdriver.common.by import By
except ImportError:
    print("Selenium non installato. Esegui:")
    print("   pip install selenium webdriver-manager")
    sys.exit(1)

try:
    from webdriver_manager.chrome import ChromeDriverManager
    USE_WDM = True
except ImportError:
    USE_WDM = False


EXPLORE_URL = "https://hub.docker.com/mcp/explore"
CATALOG_README_BASE = "https://desktop.docker.com/mcp/catalog/v3/readme"

# Categorie della pagina
CATEGORIES = [
    None,             # Pagina senza filtro
    "ai",
    "data-visualization",
    "database",
    "devops",
    "finance",
    "games",
    "communication",
    "monitoring",
    "productivity",
    "search",
]


def create_driver():
    """Crea il browser Chrome in modalita' headless."""
    opts = Options()
    opts.add_argument("--headless=new")
    opts.add_argument("--no-sandbox")
    opts.add_argument("--disable-dev-shm-usage")
    opts.add_argument("--disable-gpu")
    opts.add_argument("--window-size=1920,8000")
    opts.add_argument("--disable-blink-features=AutomationControlled")
    opts.add_argument(
        "user-agent=Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
        "AppleWebKit/537.36 (KHTML, like Gecko) Chrome/131.0.0.0 Safari/537.36"
    )

    if USE_WDM:
        service = Service(ChromeDriverManager().install())
        driver = webdriver.Chrome(service=service, options=opts)
    else:
        driver = webdriver.Chrome(options=opts)

    driver.implicitly_wait(5)
    return driver


def scroll_to_bottom(driver, max_time=45, pause=1.0):
    """Scrolla fino in fondo per caricare tutto il contenuto dinamico."""
    last_height = driver.execute_script("return document.body.scrollHeight")
    no_change = 0
    start = time.time()

    while time.time() - start < max_time:
        driver.execute_script("window.scrollTo(0, document.body.scrollHeight);")
        time.sleep(pause)
        new_height = driver.execute_script("return document.body.scrollHeight")
        if new_height == last_height:
            no_change += 1
            if no_change >= 4:
                break
        else:
            no_change = 0
        last_height = new_height


def extract_server_slugs_from_source(page_source):
    """
    Estrae gli slug dei server MCP dal sorgente HTML.
    Cerca nei riferimenti al catalogo: /catalog/v3/readme/<slug>.md
    """
    slugs = set()
    for m in re.finditer(r'/catalog/v3/(?:readme|tools)/([a-zA-Z0-9_.-]+)(?:\.md|\.json)', page_source):
        slug = m.group(1)
        slugs.add(slug)
    return slugs


def fetch_readme_and_extract_github(slug, session):
    """
    Scarica il readme di un server e ne estrae:
    - Repository URL (GitHub)
    - Author
    - Nome del server
    """
    url = f"{CATALOG_README_BASE}/{slug}.md"
    try:
        resp = session.get(url, timeout=15)
        if resp.status_code != 200:
            return None

        content = resp.text

        result = {
            "slug": slug,
            "name": "",
            "author": "",
            "repository": "",
        }

        # Estrai il nome dal titolo (prima riga # ...)
        name_match = re.search(r'^#\s+(.+)$', content, re.MULTILINE)
        if name_match:
            result["name"] = name_match.group(1).strip()

        # Estrai il Repository URL
        repo_match = re.search(r'\*\*Repository\*\*\|?\s*(https?://github\.com/[^\s\n|]+)', content)
        if repo_match:
            result["repository"] = repo_match.group(1).strip().rstrip("|").rstrip()

        # Estrai l'Author
        author_match = re.search(r'\*\*Author\*\*\|?\s*\[([^\]]+)\]', content)
        if author_match:
            result["author"] = author_match.group(1).strip()

        # Se non troviamo il repository nel formato tabella, cerchiamo altrove
        if not result["repository"]:
            gh_match = re.search(r'https?://github\.com/([a-zA-Z0-9_.-]+/[a-zA-Z0-9_.-]+)', content)
            if gh_match:
                result["repository"] = gh_match.group(0).split(")")[0].split(" ")[0].rstrip("/")

        return result if result["repository"] else None

    except requests.RequestException:
        return None


def save_to_excel(results, filename):
    """Salva i link GitHub in un file Excel nella prima colonna."""
    wb = Workbook()
    ws = wb.active
    ws.title = "MCP Servers - GitHub Links"

    # Headers
    headers = [
        ("A", "GitHub Link"),
        ("B", "Nome Server"),
        ("C", "Autore"),
        ("D", "Slug Docker"),
    ]
    for col, title in headers:
        cell = ws[f"{col}1"]
        cell.value = title
        cell.font = Font(bold=True, size=12, color="FFFFFF", name="Arial")
        cell.fill = PatternFill("solid", fgColor="4472C4")
        cell.alignment = Alignment(horizontal="center")

    # Ordina per link GitHub
    sorted_results = sorted(results, key=lambda x: x["repository"].lower())

    for i, r in enumerate(sorted_results, start=2):
        ws[f"A{i}"] = r["repository"]
        ws[f"A{i}"].font = Font(name="Arial", size=10, color="0563C1", underline="single")

        ws[f"B{i}"] = r["name"]
        ws[f"B{i}"].font = Font(name="Arial", size=10)

        ws[f"C{i}"] = r["author"]
        ws[f"C{i}"].font = Font(name="Arial", size=10)

        ws[f"D{i}"] = r["slug"]
        ws[f"D{i}"].font = Font(name="Arial", size=10)

    ws.column_dimensions["A"].width = 65
    ws.column_dimensions["B"].width = 40
    ws.column_dimensions["C"].width = 25
    ws.column_dimensions["D"].width = 35
    ws.auto_filter.ref = f"A1:D{len(sorted_results) + 1}"
    ws.freeze_panes = "A2"

    wb.save(filename)
    return len(sorted_results)


def main():
    print("=" * 65)
    print("  Docker Hub MCP Servers - GitHub Link Scraper")
    print("  Fonte: https://hub.docker.com/mcp/explore")
    print("=" * 65)
    print()

    script_dir = os.path.dirname(os.path.abspath(__file__))
    output_file = os.path.join(script_dir, "10_docker_catalog.xlsx")

    # ── FASE 1: Raccogli gli slug dei server da tutte le categorie ──
    print("[FASE 1] Raccolta nomi server da tutte le categorie...")
    print("  Avvio browser headless...")
    driver = create_driver()

    all_slugs = set()

    try:
        for cat in CATEGORIES:
            if cat is None:
                url = EXPLORE_URL
                label = "Tutti (home)"
            else:
                url = f"{EXPLORE_URL}?categories={cat}"
                label = cat

            print(f"  -> {label}...", end=" ", flush=True)
            try:
                driver.get(url)
                time.sleep(4)
                scroll_to_bottom(driver, max_time=25, pause=0.8)

                source = driver.page_source
                slugs = extract_server_slugs_from_source(source)
                new_slugs = slugs - all_slugs
                all_slugs.update(slugs)

                print(f"+{len(new_slugs)} nuovi (totale: {len(all_slugs)})")
            except Exception as e:
                print(f"errore: {e}")

    finally:
        driver.quit()
        print("  Browser chiuso.\n")

    print(f"  Totale server MCP trovati: {len(all_slugs)}")

    # ── FASE 2: Scarica i readme e estrai i link GitHub ──
    print(f"\n[FASE 2] Download readme per {len(all_slugs)} server...")

    session = requests.Session()
    session.headers.update({
        "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36"
    })

    results = []
    no_github = []
    errors = []
    sorted_slugs = sorted(all_slugs)

    for i, slug in enumerate(sorted_slugs):
        print(f"  [{i+1}/{len(sorted_slugs)}] {slug}...", end=" ", flush=True)
        result = fetch_readme_and_extract_github(slug, session)

        if result:
            results.append(result)
            print(f"-> {result['repository']}")
        else:
            no_github.append(slug)
            print("nessun link GitHub")

        # Rate limiting gentile
        time.sleep(0.3)

    # ── FASE 3: Salva in Excel ──
    if results:
        print(f"\n[FASE 3] Salvataggio in Excel...")
        count = save_to_excel(results, output_file)

        print(f"\n{'=' * 65}")
        print(f"  COMPLETATO!")
        print(f"  Link GitHub trovati: {count}")
        print(f"  Server senza GitHub: {len(no_github)}")
        print(f"  File salvato: {output_file}")
        print(f"{'=' * 65}")

        if no_github:
            print(f"\n  Server senza link GitHub:")
            for s in no_github:
                print(f"    - {s}")
    else:
        print("\nNessun link GitHub trovato.")


if __name__ == "__main__":
    main()
