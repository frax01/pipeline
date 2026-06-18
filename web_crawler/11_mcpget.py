from playwright.sync_api import sync_playwright
from openpyxl import Workbook, load_workbook
import time
import os
import sys
from datetime import datetime

BASE_URL = "https://mcp-get.com"
START_URL = "https://mcp-get.com/"
OUTPUT_XLSX = "11_mcpget.xlsx"

SCROLL_STEP = 8000
SCROLL_DELAY = 1.0

INVALID_PATHS = [
    "/getting-started", "/docs", "/about", "/blog"
]

# =========================
# Logging
# =========================

def log(msg):
    ts = datetime.now().strftime("%H:%M:%S")
    print(f"[{ts}] {msg}")
    sys.stdout.flush()

# =========================
# Safe navigation
# =========================

def safe_goto(page, url, retries=3):
    for i in range(1, retries + 1):
        try:
            page.goto(url, timeout=60000)
            page.wait_for_load_state("networkidle")
            return True
        except Exception as e:
            log(f"[WARN] goto fallito ({i}/{retries}): {e}")
            time.sleep(2)
    return False

# =========================
# Excel / Resume
# =========================

def load_or_create_workbook():
    if os.path.exists(OUTPUT_XLSX):
        wb = load_workbook(OUTPUT_XLSX)
        ws = wb.active
        processed = set()
        for row in ws.iter_rows(min_row=2, values_only=True):
            if row and row[1]:
                processed.add(row[1])
        log(f"Ripresa: {len(processed)} server già processati")
        return wb, ws, processed
    else:
        wb = Workbook()
        ws = wb.active
        ws.title = "MCP-GET Servers"
        ws.append(["Server name", "Detail page", "GitHub repo"])
        log("Creato nuovo file Excel")
        return wb, ws, set()

# =========================
# Main
# =========================

def main():
    wb, ws, processed = load_or_create_workbook()
    seen = set(processed)

    with sync_playwright() as p:
        browser = p.chromium.launch(headless=True)
        page = browser.new_page()

        try:
            log("Apro pagina principale")
            safe_goto(page, START_URL)

            last_height = 0

            while True:
                height = page.evaluate("document.body.scrollHeight")
                log(f"Altezza pagina: {height}")

                links = page.locator("a[href^='/']")
                count = links.count()
                log(f"Link visibili: {count}")

                for i in range(count):
                    a = links.nth(i)
                    href = a.get_attribute("href")

                    if not href:
                        continue
                    if href.startswith("http") or href == "/":
                        continue
                    if any(href.startswith(p) for p in INVALID_PATHS):
                        continue

                    detail_url = BASE_URL + href
                    if detail_url in seen:
                        continue

                    seen.add(detail_url)
                    name = a.inner_text().split("\n")[0].strip() or href.strip("/")

                    log(f"→ [{i}/{count}] Server: {name}")
                    log(f"   URL: {detail_url}")

                    if not safe_goto(page, detail_url):
                        log("   [SKIP] pagina irraggiungibile")
                        continue

                    # 🔥 ESTRAZIONE CORRETTA
                    github_link = "link non trovato"
                    items = page.locator("li.font-mono")
                    n = items.count()

                    if n > 0:
                        last = items.nth(n - 1)
                        link = last.locator("a[href]")
                        if link.count() > 0:
                            gh = link.first.get_attribute("href")
                            if gh and "github.com" in gh:
                                github_link = gh
                                log(f"   GitHub: {github_link}")
                    else:
                        log("   nessun link trovato nella pagina")

                    ws.append([name, detail_url, github_link])
                    wb.save(OUTPUT_XLSX)

                    safe_goto(page, START_URL)

                if height == last_height:
                    log("Fine pagina raggiunta")
                    break

                last_height = height
                page.mouse.wheel(0, SCROLL_STEP)
                time.sleep(SCROLL_DELAY)

            browser.close()
            wb.save(OUTPUT_XLSX)
            log("Completato")

        except KeyboardInterrupt:
            log("INTERRUZIONE MANUALE – stato salvato")
            wb.save(OUTPUT_XLSX)
            browser.close()

# =========================
# Entry
# =========================

if __name__ == "__main__":
    main()
