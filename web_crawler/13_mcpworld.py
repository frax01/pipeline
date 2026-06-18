"""
MCP World Scraper - Scarica tutti i link GitHub dei server MCP da mcpworld.com
Usa l'API con paginazione corretta (parametri pn/pl).

Uso:
    python scrape_mcpworld.py

Il file progress.json tiene traccia dell'avanzamento.
Se il processo viene interrotto (Ctrl+C), rilanciarlo per riprendere.

Requisiti:
    pip install requests openpyxl
"""

import json
import time
import os
import sys
import signal
import tempfile
from pathlib import Path

# === Auto-install dipendenze ===
try:
    import requests
except ImportError:
    os.system(f"{sys.executable} -m pip install requests")
    import requests

try:
    from openpyxl import Workbook
except ImportError:
    os.system(f"{sys.executable} -m pip install openpyxl")
    from openpyxl import Workbook

# === CONFIGURAZIONE ===
SCRIPT_DIR = Path(__file__).parent
EXCEL_FILE = SCRIPT_DIR / "13_mcpworld.xlsx"
PROGRESS_FILE = SCRIPT_DIR / "progress.json"
URLS_FILE = SCRIPT_DIR / "collected_urls.txt"  # Un URL per riga, più efficiente di JSON
API_URL = "https://www.mcpworld.com/api/mcp-market/servers"
PAGE_SIZE = 100          # Server per pagina
DELAY = 0.3              # Secondi tra richieste
SAVE_EVERY = 500         # Salva ogni N nuovi link
MAX_RETRIES = 3          # Tentativi per richiesta
MAX_CONSECUTIVE_EMPTY = 5  # Stop dopo N pagine vuote

HEADERS = {
    "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36",
    "Accept": "application/json",
    "Referer": "https://www.mcpworld.com/",
}

running = True

def signal_handler(sig, frame):
    global running
    print("\n⏸️  Interruzione richiesta! Salvataggio...")
    running = False

signal.signal(signal.SIGINT, signal_handler)


def load_state():
    """Carica stato: last_pn da progress.json, URLs da collected_urls.txt"""
    last_pn = -1
    urls = set()
    
    if PROGRESS_FILE.exists():
        try:
            with open(PROGRESS_FILE, "r", encoding="utf-8") as f:
                data = json.load(f)
            last_pn = data.get("last_pn", -1)
            # Migrazione: se il vecchio formato ha collected_urls, caricali
            if "collected_urls" in data and data["collected_urls"]:
                urls.update(data["collected_urls"])
        except Exception:
            pass
    
    if URLS_FILE.exists():
        with open(URLS_FILE, "r", encoding="utf-8") as f:
            for line in f:
                line = line.strip()
                if line:
                    urls.add(line)
    
    return last_pn, urls


def save_state(last_pn, urls):
    """Salva stato in modo atomico"""
    # Salva progress (solo last_pn e conteggio)
    progress = {"last_pn": last_pn, "total_urls": len(urls)}
    tmp_progress = str(PROGRESS_FILE) + ".tmp"
    try:
        with open(tmp_progress, "w", encoding="utf-8") as f:
            json.dump(progress, f)
        os.replace(tmp_progress, str(PROGRESS_FILE))
    except Exception as e:
        print(f"  ⚠️  Errore salvataggio progress: {e}")
    
    # Salva URLs (un URL per riga)
    tmp_urls = str(URLS_FILE) + ".tmp"
    try:
        with open(tmp_urls, "w", encoding="utf-8") as f:
            for url in sorted(urls):
                f.write(url + "\n")
        os.replace(tmp_urls, str(URLS_FILE))
    except Exception as e:
        print(f"  ⚠️  Errore salvataggio URLs: {e}")


def save_excel(urls):
    """Salva URLs nel file Excel"""
    try:
        wb = Workbook()
        ws = wb.active
        ws.title = "MCP Server GitHub Links"
        ws["A1"] = "GitHub URL"
        ws["A1"].font = ws["A1"].font.copy(bold=True)
        for i, url in enumerate(sorted(urls), start=2):
            ws[f"A{i}"] = url
        tmp_excel = str(EXCEL_FILE) + ".tmp"
        wb.save(tmp_excel)
        os.replace(tmp_excel, str(EXCEL_FILE))
    except Exception as e:
        print(f"  ⚠️  Errore salvataggio Excel: {e}")


def fetch_page(pn, retries=MAX_RETRIES):
    """Recupera una pagina di server dall'API"""
    params = {
        "wd": "",
        "type": "tag",
        "pn": pn,
        "lg": "en",
        "pl": PAGE_SIZE,
    }
    for attempt in range(retries):
        try:
            resp = requests.get(API_URL, params=params, headers=HEADERS, timeout=30)
            resp.raise_for_status()
            data = resp.json()
            if data.get("code") == 0:
                return data.get("data", {})
            return None
        except Exception as e:
            if attempt < retries - 1:
                time.sleep(2 ** attempt)
            else:
                print(f"  ❌ Errore pn={pn}: {e}")
                return None
    return None


def extract_github_urls(data):
    """Estrae tutti gli URL GitHub dalla risposta API"""
    github_urls = []
    for group in data.get("mcpList", []):
        for server in group.get("servers", []):
            surl = server.get("serverUrl", "")
            if surl and "github.com" in surl.lower():
                github_urls.append(surl)
    return github_urls


def main():
    global running

    print("🌐 MCP World Scraper")
    print("=" * 60)

    last_pn, urls = load_state()
    start_pn = last_pn + 1

    if urls:
        print(f"📌 Ripresa: {len(urls):,} link, da pagina {start_pn}")
    else:
        print("🆕 Nuova sessione")

    # Prima richiesta per totale
    print(f"\n🔍 Controllo totale server...")
    first_data = fetch_page(0)
    if not first_data:
        print("❌ Impossibile raggiungere l'API.")
        return
    
    total = first_data.get("count", 0)
    max_pages = (total // PAGE_SIZE) + 2
    print(f"   Totale server: {total:,}")
    print(f"   Pagine stimate: ~{max_pages}")
    print(f"   Ctrl+C per interrompere\n")

    # Se riprendiamo, non ricalcoliamo pagina 0
    if start_pn == 0:
        page_urls = extract_github_urls(first_data)
        new = 0
        for u in page_urls:
            if u not in urls:
                urls.add(u)
                new += 1
        if new > 0:
            print(f"  📡 pn=0: +{new} GitHub (tot: {len(urls):,})")
        start_pn = 1

    new_since_save = 0
    consecutive_empty = 0

    for pn in range(start_pn, max_pages + 100):
        if not running:
            break

        data = fetch_page(pn)
        if not data:
            consecutive_empty += 1
            if consecutive_empty >= MAX_CONSECUTIVE_EMPTY:
                print(f"\n⏹️  {MAX_CONSECUTIVE_EMPTY} pagine vuote. Fine.")
                break
            continue

        page_urls = extract_github_urls(data)
        if not page_urls:
            consecutive_empty += 1
            if consecutive_empty >= MAX_CONSECUTIVE_EMPTY:
                print(f"\n⏹️  {MAX_CONSECUTIVE_EMPTY} pagine senza server. Fine.")
                break
            continue

        page_new = 0
        for u in page_urls:
            if u not in urls:
                urls.add(u)
                page_new += 1

        new_since_save += page_new

        if page_new > 0:
            consecutive_empty = 0
            print(f"  📡 pn={pn}: +{page_new} GitHub (tot: {len(urls):,})")
        else:
            consecutive_empty += 1
            if pn % 100 == 0:
                print(f"  📡 pn={pn}: nessun nuovo (tot: {len(urls):,})")

        # Salva periodicamente
        if new_since_save >= SAVE_EVERY or (pn % 100 == 0 and new_since_save > 0):
            save_state(pn, urls)
            save_excel(urls)
            print(f"  💾 Salvati {len(urls):,} link (pn={pn})")
            new_since_save = 0

        time.sleep(DELAY)

    # Salvataggio finale
    save_state(pn if 'pn' in dir() else last_pn, urls)
    save_excel(urls)

    print("\n" + "=" * 60)
    print(f"🎉 Completato!")
    print(f"   📊 {len(urls):,} link GitHub unici")
    print(f"   📁 Excel: {EXCEL_FILE}")
    print("=" * 60)
    print(f"\n🔁 Per riprendere: rilancia lo script")
    print(f"🗑️  Per ricominciare: cancella {PROGRESS_FILE.name} e {URLS_FILE.name}")


if __name__ == "__main__":
    main()
