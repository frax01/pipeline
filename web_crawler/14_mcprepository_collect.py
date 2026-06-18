"""
Scraper per mcprepository.com
Cerca tutti i server MCP tramite combinazioni di ricerca e categorie.
Salva tutti gli URL in un file Excel.
"""

from playwright.sync_api import sync_playwright
import time
import string
import itertools
import os
from openpyxl import Workbook
from openpyxl.styles import Font

OUTPUT_DIR = os.path.dirname(os.path.abspath(__file__))
BASE_URL = "https://mcprepository.com"


def get_server_links(page):
    """Estrae gli URL dei server dalla pagina corrente."""
    return set(page.eval_on_selector_all(
        "a[href]",
        """els => els
            .map(e => e.getAttribute('href'))
            .filter(h => {
                const parts = h.replace(/^\\//, '').split('/');
                return parts.length === 2
                    && !h.startsWith('/category/')
                    && !h.startsWith('/search/')
                    && !/twitter|github\\.com|linkedin|discord|mailto/.test(h);
            })
        """
    ))


def main():
    print("=" * 55)
    print("  MCP Repository Scraper -> Excel")
    print("=" * 55)

    all_urls = set()

    with sync_playwright() as p:
        browser = p.chromium.launch(headless=True)
        page = browser.new_page()

        # --- FASE 1: Categorie ---
        print("\n[1/3] Recupero categorie...")
        page.goto(BASE_URL, wait_until="networkidle")
        time.sleep(2)
        cat_links = page.eval_on_selector_all(
            "a[href]",
            "els => els.map(e => e.getAttribute('href')).filter(h => h.startsWith('/category/'))"
        )
        categories = list(set(cat_links))
        print(f"      {len(categories)} categorie trovate")

        for i, cat in enumerate(categories):
            page.goto(BASE_URL + cat, wait_until="networkidle")
            time.sleep(0.5)
            links = get_server_links(page)
            new = links - all_urls
            all_urls.update(links)
            cat_name = cat.replace("/category/", "")
            if new:
                print(f"      [{i+1}/{len(categories)}] {cat_name}: +{len(new)} (totale: {len(all_urls)})")

        print(f"      Dopo categorie: {len(all_urls)} URL unici")

        # --- FASE 2: Ricerca con combinazioni 2 lettere ---
        print("\n[2/3] Ricerca combinazioni a-z x a-z (676 query)...")
        combos = [a + b for a, b in itertools.product(string.ascii_lowercase, repeat=2)]
        total = len(combos)

        for i, combo in enumerate(combos):
            try:
                page.goto(f"{BASE_URL}/search/{combo}", wait_until="networkidle", timeout=15000)
                time.sleep(0.3)
                links = get_server_links(page)
                new = links - all_urls
                all_urls.update(links)
            except Exception:
                new = set()

            if (i + 1) % 50 == 0 or new:
                status = f"+{len(new)}" if new else "0"
                print(f"      [{i+1}/{total}] '{combo}': {status} (totale: {len(all_urls)})")

        print(f"      Dopo ricerca lettere: {len(all_urls)} URL unici")

        # --- FASE 3: Ricerca con combinazioni numeriche e miste ---
        print("\n[3/3] Ricerca combinazioni extra (numeri, termini comuni)...")
        extra_queries = []
        for d in string.digits:
            for c in string.ascii_lowercase:
                extra_queries.append(d + c)
                extra_queries.append(c + d)
        extra_queries += [
            "mcp-server", "mcp-tool", "claude", "openai", "gpt", "langchain",
            "anthropic", "gemini", "ollama", "huggingface", "pytorch", "tensorflow",
            "react", "vue", "angular", "node", "python", "rust", "golang",
            "aws", "azure", "gcp", "firebase", "supabase", "prisma", "mongo",
            "redis", "elastic", "kafka", "rabbit", "graphql", "rest", "grpc",
            "slack", "notion", "jira", "confluence", "linear", "figma",
            "stripe", "twilio", "sendgrid", "vercel", "netlify", "railway",
            "browser", "scrape", "crawl", "pdf", "image", "video", "audio",
            "email", "sms", "push", "webhook", "cron", "queue", "cache",
            "auth", "oauth", "jwt", "crypto", "blockchain", "nft", "defi",
        ]
        done_combos = set(combos)
        extra_queries = [q for q in extra_queries if q not in done_combos]
        extra_queries = list(dict.fromkeys(extra_queries))

        total_extra = len(extra_queries)
        for i, query in enumerate(extra_queries):
            try:
                page.goto(f"{BASE_URL}/search/{query}", wait_until="networkidle", timeout=15000)
                time.sleep(0.3)
                links = get_server_links(page)
                new = links - all_urls
                all_urls.update(links)
            except Exception:
                new = set()

            if (i + 1) % 100 == 0 or new:
                status = f"+{len(new)}" if new else "0"
                print(f"      [{i+1}/{total_extra}] '{query}': {status} (totale: {len(all_urls)})")

        browser.close()

    print(f"\n[+] TOTALE URL UNICI: {len(all_urls)}")

    # --- Salva in Excel ---
    print("\n[*] Salvataggio Excel...")
    wb = Workbook()
    ws = wb.active
    ws.title = "MCP Servers"
    ws["A1"] = "URL"
    ws["A1"].font = Font(bold=True)
    ws.column_dimensions["A"].width = 80

    sorted_urls = sorted(all_urls)
    for i, url in enumerate(sorted_urls, start=2):
        full_url = BASE_URL + url if url.startswith("/") else url
        ws[f"A{i}"] = full_url

    excel_path = os.path.join(OUTPUT_DIR, "14_mcprepository_servers.xlsx")
    wb.save(excel_path)

    print(f"\n{'=' * 55}")
    print(f"  COMPLETATO!")
    print(f"  {len(all_urls)} URL salvati in:")
    print(f"  {excel_path}")
    print(f"{'=' * 55}")


if __name__ == "__main__":
    main()
