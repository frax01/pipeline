"""
Legge 14_mcprepository_servers.xlsx, visita ogni URL e estrae il link GitHub.
Salva il risultato in 14_mcprepository_github.xlsx.
Usa requests + BeautifulSoup (piu' veloce e affidabile di Playwright).
"""

import requests
from bs4 import BeautifulSoup
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font
from concurrent.futures import ThreadPoolExecutor, as_completed
import os
import json
import time

SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
INPUT_FILE = os.path.join(SCRIPT_DIR, "14_mcprepository_servers.xlsx")
OUTPUT_FILE = os.path.join(SCRIPT_DIR, "14_mcprepository_github.xlsx")
PROGRESS_FILE = os.path.join(SCRIPT_DIR, "_progress.json")
HEADERS = {
    "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36"
}
MAX_WORKERS = 3


def load_urls():
    wb = load_workbook(INPUT_FILE, read_only=True)
    ws = wb.active
    urls = []
    for row in ws.iter_rows(min_row=2, max_col=1, values_only=True):
        if row[0]:
            urls.append(str(row[0]).strip())
    wb.close()
    return urls


def load_progress():
    if os.path.exists(PROGRESS_FILE):
        with open(PROGRESS_FILE, "r", encoding="utf-8") as f:
            return json.load(f)
    return {}


def save_progress(results):
    with open(PROGRESS_FILE, "w", encoding="utf-8") as f:
        json.dump(results, f, ensure_ascii=False)


def save_excel(results, urls):
    wb = Workbook()
    ws = wb.active
    ws.title = "MCP Servers - GitHub"
    ws["A1"] = "MCP Repository URL"
    ws["B1"] = "GitHub URL"
    ws["A1"].font = Font(bold=True)
    ws["B1"].font = Font(bold=True)
    ws.column_dimensions["A"].width = 60
    ws.column_dimensions["B"].width = 80

    row = 2
    for url in urls:
        github_url = results.get(url, "")
        if github_url:
            ws[f"A{row}"] = url
            ws[f"B{row}"] = github_url
            row += 1

    wb.save(OUTPUT_FILE)


def extract_github_url(mcp_url):
    """Scarica la pagina e cerca il link GitHub con title='... GitHub'."""
    try:
        time.sleep(0.2)
        resp = requests.get(mcp_url, headers=HEADERS, timeout=15)
        if resp.status_code == 429:
            time.sleep(5)
            resp = requests.get(mcp_url, headers=HEADERS, timeout=15)
        resp.raise_for_status()
        soup = BeautifulSoup(resp.text, "html.parser")

        # 1) Link con title che finisce con "GitHub" (il pulsante principale)
        for a in soup.find_all("a", href=True, title=True):
            title = a.get("title", "")
            href = a["href"]
            if title.endswith("GitHub") and "github.com" in href and "github.com/mcprepository" not in href:
                return href.split("?ref=")[0]

        # 2) Fallback: link con testo "GitHub Link"
        for a in soup.find_all("a", href=True):
            text = a.get_text(strip=True)
            href = a["href"]
            if "GitHub Link" in text and "github.com" in href and "github.com/mcprepository" not in href:
                return href.split("?ref=")[0]

        return ""
    except Exception:
        return ""


def main():
    urls = load_urls()
    print(f"[*] Caricati {len(urls)} URL da {INPUT_FILE}")

    results = load_progress()
    already_done = len(results)
    if already_done > 0:
        print(f"[*] Ripristinato progresso: {already_done} gia' completati")

    todo = [u for u in urls if u not in results]
    print(f"[*] Da elaborare: {len(todo)}")
    print(f"[*] Workers paralleli: {MAX_WORKERS}")

    if not todo:
        print("[+] Tutto gia' completato!")
        save_excel(results, urls)
        print(f"[+] Salvato: {OUTPUT_FILE}")
        return

    done_count = already_done
    found_count = sum(1 for v in results.values() if v)

    with ThreadPoolExecutor(max_workers=MAX_WORKERS) as executor:
        batch_size = 200
        for batch_start in range(0, len(todo), batch_size):
            batch = todo[batch_start:batch_start + batch_size]
            futures = {executor.submit(extract_github_url, url): url for url in batch}

            for future in as_completed(futures):
                url = futures[future]
                github_url = future.result()
                results[url] = github_url
                done_count += 1

                if github_url:
                    found_count += 1
                    short = url.split(".com")[1] if ".com" in url else url
                    print(f"  [{done_count}/{len(urls)}] {short} -> {github_url}")

                if done_count % 500 == 0:
                    print(f"  [{done_count}/{len(urls)}] progresso... ({found_count} GitHub trovati)")

            save_progress(results)
            if (batch_start + batch_size) % 1000 < batch_size:
                save_excel(results, urls)
                print(f"  [*] Progresso salvato su Excel ({done_count}/{len(urls)})")

    save_excel(results, urls)

    found = sum(1 for v in results.values() if v)
    print(f"\n{'=' * 55}")
    print(f"  COMPLETATO!")
    print(f"  {found}/{len(urls)} link GitHub trovati")
    print(f"  Salvato: {OUTPUT_FILE}")
    print(f"{'=' * 55}")

    if os.path.exists(PROGRESS_FILE):
        os.remove(PROGRESS_FILE)


if __name__ == "__main__":
    main()
