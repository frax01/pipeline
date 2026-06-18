"""
Script per scaricare tutti i link GitHub degli MCP server
dal registry AIBase: https://mcp.aibase.com/explore

L'API ha un limite di ~10 pagine per query.
Strategia: combinare filtri (categorie + ordinamenti + lingue) per coprire
il massimo numero di server. Ogni combo restituisce fino a 400 risultati.

Requisiti:
    pip install requests openpyxl
"""

import time
import sys
import os
import requests
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment

API_URL = "https://mcpapi.aibase.cn/api/mcp/querypage"

HEADERS = {
    "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 "
                  "(KHTML, like Gecko) Chrome/131.0.0.0 Safari/537.36",
    "Content-Type": "application/json",
    "Accept": "application/json",
    "Origin": "https://mcp.aibase.com",
    "Referer": "https://mcp.aibase.com/explore",
}

SORTS = ["rating|desc", "rating|asc", "download|desc", "download|asc",
         "stargazersCount|desc", "stargazersCount|asc"]

CATEGORIES = [
    "", "Developer tools", "Research and data", "Knowledge management",
    "Education", "Database", "Finance", "Search tools", "Security",
    "Version control", "Cloud platform", "Image and video processing",
    "Monitoring", "Communication", "OS automation", "Entertainment",
    "Games", "Note-taking", "Marketing", "Schedule management",
    "Home automation and iot", "Location services", "Browser automation",
    "File system", "E-commerce", "Customer support", "Social media",
    "Voice processing", "Health and wellness", "Customer data platform",
    "Travel and transportation", "Virtualization", "Cloud storage",
    "Law and compliance", "Art and culture", "Language translation",
    "Data visualization", "Other", "AI chatbot",
]

DEV_LANGS = ["", "Python", "TypeScript", "JavaScript", "Go", "Rust", "Java", "C#"]

MAX_PAGES = 10


def fetch_items(session, page_no, sort="rating|desc", class_name="", dev_lang=""):
    """Scarica una pagina dall'API. Restituisce (items, total_count, error)."""
    ts = int(time.time() * 1000)
    body = {
        "certStatus": "", "className": class_name, "contentLang": "",
        "devLang": dev_lang, "langType": "en", "mcpName": "", "mcpType": 1,
        "pageNo": page_no, "pageSize": 40, "pos": "", "sort": sort, "topics": "",
    }
    try:
        r = session.post(f"{API_URL}?t={ts}&langType=en", json=body,
                        headers=HEADERS, timeout=20)
        if r.status_code == 200:
            d = r.json()
            if d.get("code") == 200:
                data = d["data"]
                return data.get("list", []), data.get("totalCount", 0), None
        return [], 0, f"HTTP {r.status_code}"
    except Exception as e:
        return [], 0, str(e)


def scrape_with_filters(session, all_repos, sort, cat, lang):
    """Pagina fino a MAX_PAGES per una combinazione di filtri."""
    added = 0
    for page in range(1, MAX_PAGES + 1):
        items, total, err = fetch_items(session, page, sort, cat, lang)
        if err or not items:
            break
        page_new = 0
        for item in items:
            fn = item.get("fullName", "")
            if fn and "/" in fn and fn not in all_repos:
                all_repos[fn] = item.get("mcpName", "")
                page_new += 1
                added += 1
        if page_new == 0 and page > 1:
            break
        time.sleep(0.1)
    return added


def save_to_excel(repos, filename):
    """Salva i link GitHub in un file Excel."""
    wb = Workbook()
    ws = wb.active
    ws.title = "AIBase MCP - GitHub Links"

    for col, title in [("A", "GitHub Link"), ("B", "Owner"),
                       ("C", "Repository"), ("D", "Server Name")]:
        cell = ws[f"{col}1"]
        cell.value = title
        cell.font = Font(bold=True, size=12, color="FFFFFF", name="Arial")
        cell.fill = PatternFill("solid", fgColor="2E86C1")
        cell.alignment = Alignment(horizontal="center")

    for i, (fn, name) in enumerate(sorted(repos.items(), key=lambda x: x[0].lower()), start=2):
        parts = fn.split("/", 1)
        ws[f"A{i}"] = f"https://github.com/{fn}"
        ws[f"A{i}"].font = Font(name="Arial", size=10, color="0563C1", underline="single")
        ws[f"B{i}"] = parts[0]
        ws[f"B{i}"].font = Font(name="Arial", size=10)
        ws[f"C{i}"] = parts[1] if len(parts) > 1 else ""
        ws[f"C{i}"].font = Font(name="Arial", size=10)
        ws[f"D{i}"] = name
        ws[f"D{i}"].font = Font(name="Arial", size=10)

    ws.column_dimensions["A"].width = 65
    ws.column_dimensions["B"].width = 30
    ws.column_dimensions["C"].width = 40
    ws.column_dimensions["D"].width = 40
    ws.auto_filter.ref = f"A1:D{len(repos) + 1}"
    ws.freeze_panes = "A2"
    wb.save(filename)
    return len(repos)


def main():
    print("=" * 65)
    print("  AIBase MCP Registry - GitHub Link Scraper")
    print("  Fonte: https://mcp.aibase.com/explore")
    print("=" * 65)

    script_dir = os.path.dirname(os.path.abspath(__file__))
    output_file = os.path.join(script_dir, "15_aibase.xlsx")

    session = requests.Session()
    all_repos = {}
    query_count = 0

    # ── FASE 1: Ordinamenti senza filtri ──
    print("\n[1/3] Ordinamenti globali...")
    for sort in SORTS:
        n = scrape_with_filters(session, all_repos, sort, "", "")
        query_count += MAX_PAGES
        if n > 0:
            print(f"  {sort}: +{n} (totale: {len(all_repos)})")

    # ── FASE 2: Categorie × ordinamenti ──
    print(f"\n[2/3] Categorie x ordinamenti...")
    for cat in CATEGORIES:
        if not cat:
            continue
        cat_new = 0
        for sort in SORTS:
            n = scrape_with_filters(session, all_repos, sort, cat, "")
            cat_new += n
            query_count += MAX_PAGES
        if cat_new > 0:
            print(f"  {cat}: +{cat_new} (totale: {len(all_repos)})")

    # ── FASE 3: Categorie × lingue × ordinamenti (solo top 3 sort) ──
    print(f"\n[3/3] Categorie x linguaggi x ordinamenti...")
    top_sorts = ["rating|desc", "rating|asc", "download|desc"]
    for cat in CATEGORIES:
        if not cat:
            continue
        for lang in DEV_LANGS:
            if not lang:
                continue
            for sort in top_sorts:
                n = scrape_with_filters(session, all_repos, sort, cat, lang)
                query_count += MAX_PAGES
                if n > 0:
                    print(f"  {cat}/{lang}/{sort}: +{n} (totale: {len(all_repos)})")

    print(f"\n  Query totali effettuate: ~{query_count}")
    print(f"  Repository GitHub unici: {len(all_repos)}")

    # ── Salva ──
    if all_repos:
        print(f"\n[SALVATAGGIO]...")
        count = save_to_excel(all_repos, output_file)
        print(f"\n{'=' * 65}")
        print(f"  COMPLETATO!")
        print(f"  Link GitHub trovati: {count}")
        print(f"  File: {output_file}")
        print(f"{'=' * 65}")
    else:
        print("\nNessun link GitHub trovato.")


if __name__ == "__main__":
    main()
