import requests
import time
import os
from openpyxl import Workbook, load_workbook

# =========================
# CONFIGURAZIONE
# =========================

OUTPUT_FILE = "17_npm.xlsx"

SEARCH_TERM = "mcp"
PAGE_SIZE = 250        # max consentito
START_OFFSET = 34650       # <-- riparti da qui se si interrompe

SLEEP_BETWEEN_SEARCH = 1.0
SLEEP_BETWEEN_PACKAGE = 0.4

MAX_RETRIES = 5

MCP_SDK_KEYWORDS = [
    "@modelcontextprotocol/sdk",
    "fastmcp",
    "modelcontextprotocol"
]

HEADERS = {
    "Accept": "application/json",
    "User-Agent": "MCP-Research/1.0"
}

SEARCH_URL = "https://registry.npmjs.org/-/v1/search"
PACKAGE_URL = "https://registry.npmjs.org/{}"

# =========================
# FUNZIONI
# =========================

def npm_search(offset):
    for attempt in range(MAX_RETRIES):
        r = requests.get(
            SEARCH_URL,
            headers=HEADERS,
            params={
                "text": SEARCH_TERM,
                "size": PAGE_SIZE,
                "from": offset
            }
        )

        if r.status_code == 429:
            wait = 2 ** attempt
            print(f"[429] search offset {offset}, attendo {wait}s")
            time.sleep(wait)
            continue

        r.raise_for_status()
        return r.json()

    raise RuntimeError("Rate limit persistente su npm search")

def npm_package(pkg_name):
    for attempt in range(MAX_RETRIES):
        r = requests.get(
            PACKAGE_URL.format(pkg_name),
            headers=HEADERS
        )

        if r.status_code == 429:
            wait = 2 ** attempt
            print(f"[429] package {pkg_name}, attendo {wait}s")
            time.sleep(wait)
            continue

        if r.status_code == 404:
            return None

        r.raise_for_status()
        return r.json()

    raise RuntimeError(f"Rate limit persistente su {pkg_name}")

def is_mcp_package(pkg_json):
    versions = pkg_json.get("versions", {})
    if not versions:
        return False

    latest = pkg_json.get("dist-tags", {}).get("latest")
    if not latest or latest not in versions:
        return False

    deps = versions[latest].get("dependencies", {})
    if not isinstance(deps, dict):
        return False

    for dep in deps.keys():
        for sdk in MCP_SDK_KEYWORDS:
            if sdk in dep:
                return True

    return False

# =========================
# EXCEL: CREA O RIPRENDI
# =========================

if os.path.exists(OUTPUT_FILE):
    wb = load_workbook(OUTPUT_FILE)
    ws = wb.active
    print("[+] Excel npm esistente trovato")

    existing = {
        row[0] for row in ws.iter_rows(min_row=2, values_only=True)
        if row[0]
    }
else:
    wb = Workbook()
    ws = wb.active
    ws.title = "npm MCP Servers"

    ws.append([
        "package_name",
        "npm_url",
        "repository"
    ])

    existing = set()
    print("[+] Nuovo Excel npm creato")

# =========================
# RACCOLTA npm
# =========================

offset = START_OFFSET

print(f"[+] Parto da offset {offset}")

while True:
    data = npm_search(offset)

    packages = data.get("objects", [])
    if not packages:
        break

    for obj in packages:
        pkg = obj["package"]["name"]

        if pkg in existing:
            continue

        print(f"[+] Analizzo {pkg}")

        pkg_json = npm_package(pkg)
        if not pkg_json:
            continue

        if not is_mcp_package(pkg_json):
            continue

        repo = ""
        repo_info = pkg_json.get("repository")
        if isinstance(repo_info, dict):
            repo = repo_info.get("url", "")
        elif isinstance(repo_info, str):
            repo = repo_info

        ws.append([
            pkg,
            f"https://www.npmjs.com/package/{pkg}",
            repo
        ])

        existing.add(pkg)
        wb.save(OUTPUT_FILE)

        time.sleep(SLEEP_BETWEEN_PACKAGE)

    offset += PAGE_SIZE
    print(f"[+] Offset successivo: {offset}")

    time.sleep(SLEEP_BETWEEN_SEARCH)

print(f"[✓] Fine. Excel npm salvato in: {OUTPUT_FILE}")
